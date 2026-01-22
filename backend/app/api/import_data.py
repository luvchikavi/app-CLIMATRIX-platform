"""
Data Import API endpoints.

Handle CSV/Excel file uploads for bulk activity import.
Supports both sync (small files) and async (large files via Arq) processing.
"""
import csv
import io
import os
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Annotated
from uuid import UUID, uuid4

import openpyxl

from arq import ArqRedis, create_pool
from arq.connections import RedisSettings
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import FileResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlmodel import select

from app.api.auth import get_current_user
from app.config import settings
from app.database import get_session
from app.models.core import User, ReportingPeriod, Organization
from app.models.emission import Activity, Emission, EmissionFactor, ConfidenceLevel, DataSource, ImportBatch, ImportBatchStatus
from app.models.jobs import ImportJob, JobStatus, JobType
from app.services.calculation import CalculationPipeline, ActivityInput
from app.services.calculation.pipeline import CalculationError
from app.services.calculation.resolver import FactorNotFoundError
from app.services.calculation.normalizer import UnitConversionError
from app.services.ai import ColumnMapper, DataValidator

router = APIRouter()

# Directory for storing uploaded files (for async processing)
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)


async def get_redis_pool() -> ArqRedis:
    """Get Redis connection pool for queuing jobs."""
    return await create_pool(RedisSettings.from_dsn(settings.redis_url))


# ============================================================================
# Schemas
# ============================================================================

class ImportRow(BaseModel):
    """A single row from import file."""
    row_number: int
    scope: int | None = None
    category_code: str | None = None
    activity_key: str | None = None
    description: str | None = None
    quantity: float | None = None
    unit: str | None = None
    activity_date: str | None = None
    errors: list[str] = []
    warnings: list[str] = []
    is_valid: bool = True


class ImportPreview(BaseModel):
    """Preview of import file before processing."""
    total_rows: int
    valid_rows: int
    invalid_rows: int
    rows: list[ImportRow]
    columns_found: list[str]
    columns_missing: list[str]


class ImportResult(BaseModel):
    """Result of import operation."""
    total_rows: int
    imported: int
    failed: int
    errors: list[dict] = []
    import_batch_id: str | None = None  # For tracking/filtering


# Required columns for import
REQUIRED_COLUMNS = ['scope', 'category_code', 'activity_key', 'quantity', 'unit']
OPTIONAL_COLUMNS = ['description', 'activity_date', 'site_id']

# Column name aliases (allow flexible headers)
COLUMN_ALIASES = {
    'scope': ['scope', 'scope_number', 'emission_scope'],
    'category_code': ['category_code', 'category', 'ghg_category'],
    'activity_key': ['activity_key', 'activity_type', 'activity', 'type'],
    'description': ['description', 'desc', 'notes', 'name'],
    'quantity': ['quantity', 'amount', 'value', 'qty'],
    'unit': ['unit', 'units', 'uom'],
    'activity_date': ['activity_date', 'date', 'activity_month', 'period'],
    'site_id': ['site_id', 'site', 'facility', 'location'],
}


def normalize_column_name(name: str) -> str | None:
    """Map column name to standard name using aliases."""
    name_lower = name.lower().strip().replace(' ', '_')
    for standard, aliases in COLUMN_ALIASES.items():
        if name_lower in aliases:
            return standard
    return None


def parse_csv_content(content: str) -> tuple[list[dict], list[str]]:
    """Parse CSV content and return rows with normalized column names."""
    reader = csv.DictReader(io.StringIO(content))

    # Map original headers to standard names
    column_mapping = {}
    found_columns = []
    for header in reader.fieldnames or []:
        standard_name = normalize_column_name(header)
        if standard_name:
            column_mapping[header] = standard_name
            found_columns.append(standard_name)

    rows = []
    for row in reader:
        normalized_row = {}
        for original, value in row.items():
            standard = column_mapping.get(original)
            if standard:
                normalized_row[standard] = value.strip() if value else None
        rows.append(normalized_row)

    return rows, found_columns


def parse_excel_content(content: bytes) -> tuple[list[dict], list[str]]:
    """Parse Excel content and return rows with normalized column names."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    if not ws or ws.max_row < 2:
        return [], []

    # Get headers from first row
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value) if cell.value else "")

    # Map headers to standard names
    column_mapping = {}
    found_columns = []
    for i, header in enumerate(headers):
        standard_name = normalize_column_name(header)
        if standard_name:
            column_mapping[i] = standard_name
            found_columns.append(standard_name)

    # Parse data rows
    rows = []
    for row_num in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx, cell in enumerate(ws[row_num]):
            if col_idx in column_mapping:
                value = cell.value
                if value is not None:
                    # Handle dates from Excel
                    if isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d')
                    else:
                        value = str(value).strip()
                row_data[column_mapping[col_idx]] = value if value else None

        # Only add row if it has some data
        if any(row_data.values()):
            rows.append(row_data)

    wb.close()
    return rows, found_columns


def parse_file_content(content: bytes, filename: str) -> tuple[list[dict], list[str]]:
    """Parse file content based on file type."""
    filename_lower = filename.lower()

    if filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls'):
        return parse_excel_content(content)
    else:  # CSV
        try:
            text_content = content.decode('utf-8')
        except UnicodeDecodeError:
            text_content = content.decode('latin-1')
        return parse_csv_content(text_content)


def validate_row(row: dict, row_number: int, valid_activity_keys: set) -> ImportRow:
    """Validate a single row and return ImportRow with errors/warnings."""
    errors = []
    warnings = []

    # Parse scope
    scope = None
    if row.get('scope'):
        try:
            scope = int(row['scope'])
            if scope not in [1, 2, 3]:
                errors.append(f"Scope must be 1, 2, or 3 (got {scope})")
        except ValueError:
            errors.append(f"Invalid scope value: {row['scope']}")
    else:
        errors.append("Missing scope")

    # Parse category_code
    category_code = row.get('category_code')
    if not category_code:
        errors.append("Missing category_code")

    # Validate activity_key
    activity_key = row.get('activity_key')
    if not activity_key:
        errors.append("Missing activity_key")
    elif valid_activity_keys and activity_key not in valid_activity_keys:
        errors.append(f"Unknown activity_key: {activity_key}")

    # Parse quantity
    quantity = None
    if row.get('quantity'):
        try:
            quantity = float(row['quantity'].replace(',', ''))
            if quantity <= 0:
                warnings.append("Quantity is zero or negative")
        except ValueError:
            errors.append(f"Invalid quantity: {row['quantity']}")
    else:
        errors.append("Missing quantity")

    # Validate unit
    unit = row.get('unit')
    if not unit:
        errors.append("Missing unit")

    # Parse activity_date (optional, default to today)
    activity_date = row.get('activity_date')
    if activity_date:
        # Try common date formats
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d']:
            try:
                datetime.strptime(activity_date, fmt)
                break
            except ValueError:
                continue
        else:
            warnings.append(f"Could not parse date: {activity_date}, using today")
            activity_date = date.today().isoformat()
    else:
        activity_date = date.today().isoformat()

    # Description (optional)
    description = row.get('description') or f"{activity_key} import"

    return ImportRow(
        row_number=row_number,
        scope=scope,
        category_code=category_code,
        activity_key=activity_key,
        description=description,
        quantity=quantity,
        unit=unit,
        activity_date=activity_date,
        errors=errors,
        warnings=warnings,
        is_valid=len(errors) == 0,
    )


# ============================================================================
# Endpoints
# ============================================================================

@router.post("/periods/{period_id}/import/preview", response_model=ImportPreview)
async def preview_import(
    period_id: UUID,
    file: UploadFile = File(...),
    session: Annotated[AsyncSession, Depends(get_session)] = None,
    current_user: Annotated[User, Depends(get_current_user)] = None,
):
    """
    Preview import file before processing.

    Validates structure and data, returns preview with errors/warnings.
    Does NOT save anything to database.
    """
    # Verify period access
    period_query = select(ReportingPeriod).where(
        ReportingPeriod.id == period_id,
        ReportingPeriod.organization_id == current_user.organization_id,
    )
    period_result = await session.execute(period_query)
    period = period_result.scalar_one_or_none()
    if not period:
        raise HTTPException(status_code=404, detail="Reporting period not found")
    if period.is_locked:
        raise HTTPException(status_code=400, detail="Cannot import to locked period")

    # Check file type
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    filename_lower = file.filename.lower()
    if not (filename_lower.endswith('.csv') or filename_lower.endswith('.xlsx')):
        raise HTTPException(
            status_code=400,
            detail="Only CSV and Excel (.xlsx) files are supported"
        )

    # Read file content
    content = await file.read()

    # Parse file content (CSV or Excel)
    try:
        rows, found_columns = parse_file_content(content, file.filename)
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Failed to parse file: {str(e)}"
        )

    # Get valid activity keys
    factors_query = select(EmissionFactor.activity_key).where(EmissionFactor.is_active == True)
    factors_result = await session.execute(factors_query)
    valid_keys = set(factors_result.scalars().all())

    # Validate rows
    validated_rows = []
    valid_count = 0
    for i, row in enumerate(rows, start=2):  # Start at 2 (row 1 is header)
        validated = validate_row(row, i, valid_keys)
        validated_rows.append(validated)
        if validated.is_valid:
            valid_count += 1

    # Check for missing required columns
    missing_columns = [col for col in REQUIRED_COLUMNS if col not in found_columns]

    return ImportPreview(
        total_rows=len(rows),
        valid_rows=valid_count,
        invalid_rows=len(rows) - valid_count,
        rows=validated_rows[:100],  # Limit preview to first 100 rows
        columns_found=found_columns,
        columns_missing=missing_columns,
    )


@router.post("/periods/{period_id}/import", response_model=ImportResult)
async def import_activities(
    period_id: UUID,
    file: UploadFile = File(...),
    session: Annotated[AsyncSession, Depends(get_session)] = None,
    current_user: Annotated[User, Depends(get_current_user)] = None,
):
    """
    Import activities from CSV file.

    Validates and creates activities, calculating emissions for each.
    """
    # Verify period access
    period_query = select(ReportingPeriod).where(
        ReportingPeriod.id == period_id,
        ReportingPeriod.organization_id == current_user.organization_id,
    )
    period_result = await session.execute(period_query)
    period = period_result.scalar_one_or_none()
    if not period:
        raise HTTPException(status_code=404, detail="Reporting period not found")
    if period.is_locked:
        raise HTTPException(status_code=400, detail="Cannot import to locked period")

    # Get organization region
    org_query = select(Organization).where(Organization.id == current_user.organization_id)
    org_result = await session.execute(org_query)
    org = org_result.scalar_one_or_none()
    org_region = org.default_region if org and org.default_region else "Global"

    # Check file type
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    filename_lower = file.filename.lower()
    if not (filename_lower.endswith('.csv') or filename_lower.endswith('.xlsx')):
        raise HTTPException(status_code=400, detail="Only CSV and Excel (.xlsx) files are supported")

    # Read and parse file
    content = await file.read()
    try:
        rows, found_columns = parse_file_content(content, file.filename)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    # Get valid activity keys
    factors_query = select(EmissionFactor.activity_key).where(EmissionFactor.is_active == True)
    factors_result = await session.execute(factors_query)
    valid_keys = set(factors_result.scalars().all())

    # Create import batch for tracking
    import_batch = ImportBatch(
        organization_id=current_user.organization_id,
        reporting_period_id=period_id,
        file_name=file.filename,
        file_type="excel" if filename_lower.endswith('.xlsx') else "csv",
        file_size_bytes=len(content),
        status=ImportBatchStatus.PROCESSING,
        total_rows=len(rows),
        uploaded_by=current_user.id,
    )
    session.add(import_batch)
    await session.flush()  # Get the batch ID

    # Process rows
    pipeline = CalculationPipeline(session)
    imported = 0
    failed = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        validated = validate_row(row, i, valid_keys)

        if not validated.is_valid:
            failed += 1
            errors.append({
                "row": i,
                "errors": validated.errors,
            })
            continue

        try:
            # Calculate emissions
            calc_result = await pipeline.calculate(ActivityInput(
                activity_key=validated.activity_key,
                quantity=Decimal(str(validated.quantity)),
                unit=validated.unit,
                scope=validated.scope,
                category_code=validated.category_code,
                region=org_region,
                year=2024,
            ))

            # Parse date
            activity_date = date.today()
            if validated.activity_date:
                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
                    try:
                        activity_date = datetime.strptime(validated.activity_date, fmt).date()
                        break
                    except ValueError:
                        continue

            # Create activity
            activity = Activity(
                organization_id=current_user.organization_id,
                reporting_period_id=period_id,
                scope=validated.scope,
                category_code=validated.category_code,
                activity_key=validated.activity_key,
                description=validated.description,
                quantity=Decimal(str(validated.quantity)),
                unit=validated.unit,
                activity_date=activity_date,
                created_by=current_user.id,
                data_source=DataSource.IMPORT,
                import_batch_id=import_batch.id,
            )
            session.add(activity)
            await session.flush()

            # Create emission
            emission = Emission(
                activity_id=activity.id,
                emission_factor_id=calc_result.emission_factor_id,
                co2e_kg=calc_result.co2e_kg,
                co2_kg=calc_result.co2_kg,
                ch4_kg=calc_result.ch4_kg,
                n2o_kg=calc_result.n2o_kg,
                wtt_co2e_kg=calc_result.wtt_co2e_kg,
                converted_quantity=calc_result.converted_quantity,
                converted_unit=calc_result.converted_unit,
                formula=calc_result.formula,
                confidence=ConfidenceLevel(calc_result.confidence),
                resolution_strategy=calc_result.resolution_strategy,
            )
            session.add(emission)

            imported += 1

        except (FactorNotFoundError, UnitConversionError, CalculationError) as e:
            failed += 1
            errors.append({
                "row": i,
                "activity_key": validated.activity_key,
                "errors": [str(e)],
            })
        except Exception as e:
            failed += 1
            errors.append({
                "row": i,
                "errors": [f"Unexpected error: {str(e)}"],
            })

    # Update batch status
    import_batch.successful_rows = imported
    import_batch.failed_rows = failed
    import_batch.status = ImportBatchStatus.COMPLETED if failed == 0 else ImportBatchStatus.PARTIAL
    import_batch.completed_at = datetime.utcnow()
    if errors:
        import_batch.row_errors = errors[:50]

    await session.commit()

    return ImportResult(
        total_rows=len(rows),
        imported=imported,
        failed=failed,
        errors=errors[:50],  # Limit errors in response
        import_batch_id=str(import_batch.id),
    )


# ============================================================================
# Import Batch Tracking Endpoints
# ============================================================================

class ImportBatchResponse(BaseModel):
    """Response for import batch listing."""
    id: str
    file_name: str
    file_type: str
    status: str
    total_rows: int
    successful_rows: int
    failed_rows: int
    uploaded_at: datetime
    completed_at: datetime | None = None


@router.get("/import/batches", response_model=list[ImportBatchResponse])
async def list_import_batches(
    period_id: UUID = Query(default=None, description="Filter by reporting period"),
    limit: int = Query(default=20, le=100),
    session: Annotated[AsyncSession, Depends(get_session)] = None,
    current_user: Annotated[User, Depends(get_current_user)] = None,
):
    """
    List recent import batches for the organization.

    Returns a list of file uploads with their status and row counts.
    Use this to see which files have been imported and their success rate.
    """
    query = select(ImportBatch).where(
        ImportBatch.organization_id == current_user.organization_id
    )

    if period_id:
        query = query.where(ImportBatch.reporting_period_id == period_id)

    query = query.order_by(ImportBatch.uploaded_at.desc()).limit(limit)

    result = await session.execute(query)
    batches = result.scalars().all()

    return [
        ImportBatchResponse(
            id=str(batch.id),
            file_name=batch.file_name,
            file_type=batch.file_type,
            status=batch.status.value,
            total_rows=batch.total_rows,
            successful_rows=batch.successful_rows,
            failed_rows=batch.failed_rows,
            uploaded_at=batch.uploaded_at,
            completed_at=batch.completed_at,
        )
        for batch in batches
    ]


@router.get("/import/batches/{batch_id}")
async def get_import_batch(
    batch_id: UUID,
    session: Annotated[AsyncSession, Depends(get_session)] = None,
    current_user: Annotated[User, Depends(get_current_user)] = None,
):
    """
    Get details of a specific import batch including any errors.
    """
    query = select(ImportBatch).where(
        ImportBatch.id == batch_id,
        ImportBatch.organization_id == current_user.organization_id,
    )
    result = await session.execute(query)
    batch = result.scalar_one_or_none()

    if not batch:
        raise HTTPException(status_code=404, detail="Import batch not found")

    return {
        "id": str(batch.id),
        "file_name": batch.file_name,
        "file_type": batch.file_type,
        "file_size_bytes": batch.file_size_bytes,
        "status": batch.status.value,
        "total_rows": batch.total_rows,
        "successful_rows": batch.successful_rows,
        "failed_rows": batch.failed_rows,
        "skipped_rows": batch.skipped_rows,
        "error_message": batch.error_message,
        "row_errors": batch.row_errors,
        "uploaded_at": batch.uploaded_at,
        "completed_at": batch.completed_at,
    }


@router.get("/import/batches/{batch_id}/activities")
async def get_batch_activities(
    batch_id: UUID,
    session: Annotated[AsyncSession, Depends(get_session)] = None,
    current_user: Annotated[User, Depends(get_current_user)] = None,
):
    """
    Get all activities from a specific import batch.

    Use this to review what was imported from a specific file.
    """
    # Verify batch exists and belongs to user's org
    batch_query = select(ImportBatch).where(
        ImportBatch.id == batch_id,
        ImportBatch.organization_id == current_user.organization_id,
    )
    batch_result = await session.execute(batch_query)
    batch = batch_result.scalar_one_or_none()

    if not batch:
        raise HTTPException(status_code=404, detail="Import batch not found")

    # Get activities for this batch
    activities_query = select(Activity).where(
        Activity.import_batch_id == batch_id
    ).order_by(Activity.created_at)

    activities_result = await session.execute(activities_query)
    activities = activities_result.scalars().all()

    # Get emissions and factors for each activity
    result_activities = []
    for a in activities:
        # Get emission
        emission_query = select(Emission).where(Emission.activity_id == a.id)
        emission_result = await session.execute(emission_query)
        emission = emission_result.scalar_one_or_none()

        # Get factor if emission exists
        factor = None
        if emission and emission.emission_factor_id:
            factor_query = select(EmissionFactor).where(EmissionFactor.id == emission.emission_factor_id)
            factor_result = await session.execute(factor_query)
            factor = factor_result.scalar_one_or_none()

        result_activities.append({
            "id": str(a.id),
            "scope": a.scope,
            "category_code": a.category_code,
            "activity_key": a.activity_key,
            "description": a.description,
            "quantity": float(a.quantity),
            "unit": a.unit,
            "activity_date": a.activity_date.isoformat() if a.activity_date else None,
            "emission": {
                "co2e_kg": float(emission.co2e_kg) if emission else None,
                "factor_value": float(factor.co2e_factor) if factor and factor.co2e_factor else None,
                "factor_unit": factor.factor_unit if factor else None,
                "factor_source": factor.source if factor else None,
                "formula": emission.formula if emission else None,
            } if emission else None,
        })

    return {
        "batch_id": str(batch_id),
        "file_name": batch.file_name,
        "activity_count": len(activities),
        "activities": result_activities,
    }


@router.delete("/import/batches/{batch_id}")
async def delete_import_batch(
    batch_id: UUID,
    delete_activities: bool = Query(default=False, description="Also delete imported activities"),
    session: Annotated[AsyncSession, Depends(get_session)] = None,
    current_user: Annotated[User, Depends(get_current_user)] = None,
):
    """
    Delete an import batch record.

    If delete_activities=True, also deletes all activities imported in this batch.
    This is useful for undoing a bad import.
    """
    # Verify batch exists and belongs to user's org
    batch_query = select(ImportBatch).where(
        ImportBatch.id == batch_id,
        ImportBatch.organization_id == current_user.organization_id,
    )
    batch_result = await session.execute(batch_query)
    batch = batch_result.scalar_one_or_none()

    if not batch:
        raise HTTPException(status_code=404, detail="Import batch not found")

    deleted_activities = 0

    if delete_activities:
        # Delete emissions first (FK constraint)
        from sqlmodel import delete
        activities_query = select(Activity.id).where(Activity.import_batch_id == batch_id)
        activities_result = await session.execute(activities_query)
        activity_ids = [a for a in activities_result.scalars().all()]

        if activity_ids:
            await session.execute(
                delete(Emission).where(Emission.activity_id.in_(activity_ids))
            )
            delete_result = await session.execute(
                delete(Activity).where(Activity.import_batch_id == batch_id)
            )
            deleted_activities = delete_result.rowcount

    # Delete the batch
    await session.delete(batch)
    await session.commit()

    return {
        "message": "Import batch deleted",
        "batch_id": str(batch_id),
        "deleted_activities": deleted_activities,
    }


@router.get("/import/template")
async def get_import_template(
    scope: str = Query("1-2", description="Template scope: '1-2' for Scope 1&2, '3' for Scope 3, 'csv' for simple CSV")
):
    """
    Get import template file.

    - scope='1-2': Excel template for Scope 1 & 2 emissions (stationary, mobile, fugitive, electricity, etc.)
    - scope='3': Excel template for Scope 3 emissions (purchased goods, travel, waste, etc.)
    - scope='csv': Simple CSV template for quick imports

    Returns downloadable file.
    """
    import os

    base_path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    if scope == "3":
        # Scope 3 Excel template
        file_path = os.path.join(base_path, "climatrix_files", "climatrix_scope3_template_v1.xlsx")
        filename = "climatrix_scope3_template.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif scope == "1-2":
        # Scope 1 & 2 Excel template
        file_path = os.path.join(base_path, "climatrix_files", "climatrix_import_template_scope1and2_v3.xlsx")
        filename = "climatrix_scope1and2_template.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        # Simple CSV template
        template = """scope,category_code,activity_key,description,quantity,unit,activity_date
1,1.1,natural_gas_volume,Office heating - January,1500,m3,2024-01-31
1,1.1,diesel_liters,Generator fuel,200,liters,2024-01-31
2,2,electricity_il,Office electricity,10000,kWh,2024-01-31
3,3.6,flight_short_economy,Business trip TLV-London,3000,km,2024-01-15
3,3.5,waste_landfill_mixed,Office waste,500,kg,2024-01-31
"""
        return {
            "filename": "climatrix_import_template.csv",
            "content": template,
            "columns": {
                "scope": "1, 2, or 3",
                "category_code": "GHG category (e.g., 1.1, 2, 3.6)",
                "activity_key": "Exact activity identifier from reference data",
                "description": "Optional description",
                "quantity": "Numeric amount",
                "unit": "Unit of measurement (must match activity_key)",
                "activity_date": "Date in YYYY-MM-DD format",
            },
        }

    # Check if file exists
    if not os.path.exists(file_path):
        raise HTTPException(
            status_code=404,
            detail=f"Template file not found: {filename}"
        )

    return FileResponse(
        path=file_path,
        filename=filename,
        media_type=media_type
    )


# ============================================================================
# Async Import Endpoints (for large files)
# ============================================================================

class AsyncImportResponse(BaseModel):
    """Response for async import request."""
    job_id: str
    status: str
    message: str


class JobStatusResponse(BaseModel):
    """Response for job status check."""
    job_id: str
    status: str
    progress_percent: int
    total_rows: int | None
    processed_rows: int
    successful_rows: int
    failed_rows: int
    error_message: str | None
    created_at: datetime
    started_at: datetime | None
    completed_at: datetime | None
    summary: dict | None


@router.post("/periods/{period_id}/import/async", response_model=AsyncImportResponse)
async def import_activities_async(
    period_id: UUID,
    file: UploadFile = File(...),
    session: Annotated[AsyncSession, Depends(get_session)] = None,
    current_user: Annotated[User, Depends(get_current_user)] = None,
):
    """
    Import activities asynchronously (for large files).

    Creates a background job that processes the file.
    Returns immediately with a job_id for status tracking.

    Use this endpoint for files with 100+ rows.
    """
    # Verify period access
    period_query = select(ReportingPeriod).where(
        ReportingPeriod.id == period_id,
        ReportingPeriod.organization_id == current_user.organization_id,
    )
    period_result = await session.execute(period_query)
    period = period_result.scalar_one_or_none()
    if not period:
        raise HTTPException(status_code=404, detail="Reporting period not found")
    if period.is_locked:
        raise HTTPException(status_code=400, detail="Cannot import to locked period")

    # Check file type
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    filename_lower = file.filename.lower()
    if not (filename_lower.endswith('.csv') or filename_lower.endswith('.xlsx')):
        raise HTTPException(
            status_code=400,
            detail="Only CSV and Excel (.xlsx) files are supported"
        )

    # Read file content
    content = await file.read()
    file_size = len(content)

    # Save file to disk for async processing
    job_id = uuid4()
    file_ext = '.xlsx' if filename_lower.endswith('.xlsx') else '.csv'
    file_path = os.path.join(UPLOAD_DIR, f"{job_id}{file_ext}")
    with open(file_path, 'wb') as f:
        f.write(content)

    # Count rows for progress tracking
    if filename_lower.endswith('.xlsx'):
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        row_count = ws.max_row - 1 if ws else 0  # Minus header
        wb.close()
    else:
        try:
            text_content = content.decode('utf-8')
        except UnicodeDecodeError:
            text_content = content.decode('latin-1')
        row_count = len(text_content.strip().split('\n')) - 1  # Minus header

    # Create job record
    job = ImportJob(
        id=job_id,
        organization_id=current_user.organization_id,
        reporting_period_id=period_id,
        created_by=current_user.id,
        job_type=JobType.IMPORT_CSV,
        status=JobStatus.PENDING,
        original_filename=file.filename,
        file_path=file_path,
        file_size_bytes=file_size,
        total_rows=row_count,
    )
    session.add(job)
    await session.commit()

    # Queue job for async processing
    try:
        redis = await get_redis_pool()
        await redis.enqueue_job('process_import_job', str(job_id))
        await redis.close()
    except Exception as e:
        # If Redis fails, mark job as failed but don't crash
        job.mark_failed(f"Failed to queue job: {str(e)}")
        await session.commit()
        raise HTTPException(
            status_code=503,
            detail="Background worker unavailable. Please try sync import or retry later."
        )

    return AsyncImportResponse(
        job_id=str(job_id),
        status="queued",
        message=f"Import job queued. {row_count} rows will be processed in background."
    )


@router.get("/import/jobs/{job_id}", response_model=JobStatusResponse)
async def get_job_status(
    job_id: UUID,
    session: Annotated[AsyncSession, Depends(get_session)] = None,
    current_user: Annotated[User, Depends(get_current_user)] = None,
):
    """
    Get status of an import job.

    Use this to poll for progress on async imports.
    """
    # Get job (with org check for security)
    job_query = select(ImportJob).where(
        ImportJob.id == job_id,
        ImportJob.organization_id == current_user.organization_id,
    )
    job_result = await session.execute(job_query)
    job = job_result.scalar_one_or_none()

    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    return JobStatusResponse(
        job_id=str(job.id),
        status=job.status.value,
        progress_percent=job.progress_percent,
        total_rows=job.total_rows,
        processed_rows=job.processed_rows,
        successful_rows=job.successful_rows,
        failed_rows=job.failed_rows,
        error_message=job.error_message,
        created_at=job.created_at,
        started_at=job.started_at,
        completed_at=job.completed_at,
        summary=job.summary,
    )


@router.get("/import/jobs", response_model=list[JobStatusResponse])
async def list_import_jobs(
    limit: int = Query(default=10, le=100),
    session: Annotated[AsyncSession, Depends(get_session)] = None,
    current_user: Annotated[User, Depends(get_current_user)] = None,
):
    """
    List recent import jobs for the organization.
    """
    jobs_query = select(ImportJob).where(
        ImportJob.organization_id == current_user.organization_id
    ).order_by(ImportJob.created_at.desc()).limit(limit)

    jobs_result = await session.execute(jobs_query)
    jobs = jobs_result.scalars().all()

    return [
        JobStatusResponse(
            job_id=str(job.id),
            status=job.status.value,
            progress_percent=job.progress_percent,
            total_rows=job.total_rows,
            processed_rows=job.processed_rows,
            successful_rows=job.successful_rows,
            failed_rows=job.failed_rows,
            error_message=job.error_message,
            created_at=job.created_at,
            started_at=job.started_at,
            completed_at=job.completed_at,
            summary=job.summary,
        )
        for job in jobs
    ]


@router.delete("/import/jobs/{job_id}")
async def cancel_import_job(
    job_id: UUID,
    session: Annotated[AsyncSession, Depends(get_session)] = None,
    current_user: Annotated[User, Depends(get_current_user)] = None,
):
    """
    Cancel a pending import job.

    Can only cancel jobs that haven't started processing yet.
    """
    # Get job
    job_query = select(ImportJob).where(
        ImportJob.id == job_id,
        ImportJob.organization_id == current_user.organization_id,
    )
    job_result = await session.execute(job_query)
    job = job_result.scalar_one_or_none()

    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    if job.status != JobStatus.PENDING:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot cancel job with status '{job.status.value}'"
        )

    job.status = JobStatus.CANCELLED
    job.completed_at = datetime.utcnow()
    await session.commit()

    # Clean up file
    if os.path.exists(job.file_path):
        os.remove(job.file_path)

    return {"message": "Job cancelled", "job_id": str(job_id)}


# ============================================================================
# AI-Powered Smart Import Endpoints
# ============================================================================

class ColumnMappingResponse(BaseModel):
    """Response from AI column analysis."""
    success: bool
    detected_structure: str
    date_column: str | None
    quantity_column: str | None
    unit_column: str | None
    description_column: str | None
    mappings: list[dict]
    warnings: list[str]
    ai_notes: str | None


class SmartImportResponse(BaseModel):
    """Response from smart import."""
    job_id: str
    status: str
    message: str
    ai_mapping_preview: dict | None


@router.post("/periods/{period_id}/import/analyze-columns", response_model=ColumnMappingResponse)
async def analyze_import_columns(
    period_id: UUID,
    file: UploadFile = File(...),
    session: Annotated[AsyncSession, Depends(get_session)] = None,
    current_user: Annotated[User, Depends(get_current_user)] = None,
):
    """
    Analyze file columns using AI before import.

    Uses Claude AI to intelligently map columns to activity types,
    even if column names don't match the standard template.

    Returns mapping suggestions that can be confirmed before import.

    This is useful for:
    - Files with non-standard column names
    - Files in different languages
    - Files with combined data (e.g., "Natural Gas (m³)")
    """
    # Verify period access
    period_query = select(ReportingPeriod).where(
        ReportingPeriod.id == period_id,
        ReportingPeriod.organization_id == current_user.organization_id,
    )
    period_result = await session.execute(period_query)
    period = period_result.scalar_one_or_none()
    if not period:
        raise HTTPException(status_code=404, detail="Reporting period not found")

    # Check file type
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    filename_lower = file.filename.lower()
    if not (filename_lower.endswith('.csv') or filename_lower.endswith('.xlsx')):
        raise HTTPException(
            status_code=400,
            detail="Only CSV and Excel (.xlsx) files are supported"
        )

    # Read file content
    content = await file.read()

    # Parse file to get headers and sample data
    if filename_lower.endswith('.xlsx'):
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        headers = [str(cell.value) if cell.value else "" for cell in ws[1]] if ws else []
        rows = []
        if ws:
            for row_num in range(2, min(ws.max_row + 1, 7)):  # First 5 data rows
                row_values = [str(cell.value) if cell.value else "" for cell in ws[row_num]]
                if any(row_values):
                    rows.append(dict(zip(headers, row_values)))
        wb.close()
        sample_data = [list(row.values()) for row in rows]
    else:
        try:
            text_content = content.decode('utf-8')
        except UnicodeDecodeError:
            text_content = content.decode('latin-1')
        reader = csv.DictReader(io.StringIO(text_content))
        headers = list(reader.fieldnames) if reader.fieldnames else []
        rows = list(reader)
        sample_data = [list(row.values()) for row in rows[:5]]

    # Use AI Column Mapper
    mapper = ColumnMapper()
    result = mapper.map_columns(headers, sample_data)

    return ColumnMappingResponse(
        success=result.success,
        detected_structure=result.detected_structure,
        date_column=result.date_column,
        quantity_column=result.quantity_column,
        unit_column=result.unit_column,
        description_column=result.description_column,
        mappings=[
            {
                "original_header": m.original_header,
                "activity_key": m.activity_key,
                "scope": m.scope,
                "category_code": m.category_code,
                "detected_unit": m.detected_unit,
                "column_type": m.column_type,
                "confidence": m.confidence,
                "notes": m.notes,
            }
            for m in result.mappings
        ],
        warnings=result.warnings,
        ai_notes=result.ai_notes,
    )


@router.post("/periods/{period_id}/import/smart", response_model=SmartImportResponse)
async def smart_import_activities(
    period_id: UUID,
    file: UploadFile = File(...),
    session: Annotated[AsyncSession, Depends(get_session)] = None,
    current_user: Annotated[User, Depends(get_current_user)] = None,
):
    """
    Smart import using AI for intelligent column mapping.

    Unlike standard import, this endpoint:
    - Auto-detects column meanings using Claude AI
    - Maps columns to activity types intelligently
    - Handles non-standard file formats
    - Validates data and flags issues

    Best for files that don't follow the standard template.
    For large files (100+ rows), this runs asynchronously.
    """
    # Verify period access
    period_query = select(ReportingPeriod).where(
        ReportingPeriod.id == period_id,
        ReportingPeriod.organization_id == current_user.organization_id,
    )
    period_result = await session.execute(period_query)
    period = period_result.scalar_one_or_none()
    if not period:
        raise HTTPException(status_code=404, detail="Reporting period not found")
    if period.is_locked:
        raise HTTPException(status_code=400, detail="Cannot import to locked period")

    # Check file type
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    filename_lower = file.filename.lower()
    if not (filename_lower.endswith('.csv') or filename_lower.endswith('.xlsx')):
        raise HTTPException(
            status_code=400,
            detail="Only CSV and Excel (.xlsx) files are supported"
        )

    # Read file content
    content = await file.read()
    file_size = len(content)

    # Parse file and count rows
    if filename_lower.endswith('.xlsx'):
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        headers = [str(cell.value) if cell.value else "" for cell in ws[1]] if ws else []
        row_count = ws.max_row - 1 if ws else 0
        rows = []
        if ws:
            for row_num in range(2, min(ws.max_row + 1, 7)):  # First 5 data rows
                row_values = [str(cell.value) if cell.value else "" for cell in ws[row_num]]
                if any(row_values):
                    rows.append(dict(zip(headers, row_values)))
        wb.close()
        sample_data = [list(row.values()) for row in rows]
    else:
        try:
            text_content = content.decode('utf-8')
        except UnicodeDecodeError:
            text_content = content.decode('latin-1')
        row_count = len(text_content.strip().split('\n')) - 1
        reader = csv.DictReader(io.StringIO(text_content))
        headers = list(reader.fieldnames) if reader.fieldnames else []
        rows = list(reader)
        sample_data = [list(row.values()) for row in rows[:5]]

    mapper = ColumnMapper()
    mapping_result = mapper.map_columns(headers, sample_data)

    # If no activity columns detected, fail early
    activity_mappings = [m for m in mapping_result.mappings if m.column_type == "activity" and m.activity_key]
    if not activity_mappings:
        raise HTTPException(
            status_code=400,
            detail="AI could not detect any emission activity columns. "
                   "Please use the standard import template or rename columns."
        )

    # Save file for async processing
    job_id = uuid4()
    file_path = os.path.join(UPLOAD_DIR, f"{job_id}.csv")
    with open(file_path, 'wb') as f:
        f.write(content)

    # Create job record
    job = ImportJob(
        id=job_id,
        organization_id=current_user.organization_id,
        reporting_period_id=period_id,
        created_by=current_user.id,
        job_type=JobType.IMPORT_CSV,
        status=JobStatus.PENDING,
        original_filename=file.filename,
        file_path=file_path,
        file_size_bytes=file_size,
        total_rows=row_count,
        metadata={
            "smart_import": True,
            "ai_mapping_preview": {
                "detected_structure": mapping_result.detected_structure,
                "activity_columns": [
                    {"header": m.original_header, "activity_key": m.activity_key}
                    for m in activity_mappings
                ],
            }
        }
    )
    session.add(job)
    await session.commit()

    # Queue smart import job
    try:
        redis = await get_redis_pool()
        await redis.enqueue_job('smart_import_job', str(job_id))
        await redis.close()
    except Exception as e:
        job.mark_failed(f"Failed to queue job: {str(e)}")
        await session.commit()
        raise HTTPException(
            status_code=503,
            detail="Background worker unavailable. Please try standard import."
        )

    return SmartImportResponse(
        job_id=str(job_id),
        status="queued",
        message=f"Smart import queued. AI detected {len(activity_mappings)} activity columns "
                f"in {row_count} rows.",
        ai_mapping_preview={
            "detected_structure": mapping_result.detected_structure,
            "detected_columns": [
                {
                    "header": m.original_header,
                    "maps_to": m.activity_key,
                    "unit": m.detected_unit,
                    "confidence": f"{m.confidence:.0%}",
                }
                for m in activity_mappings
            ],
            "date_column": mapping_result.date_column,
            "warnings": mapping_result.warnings[:3],  # First 3 warnings
        }
    )


@router.post("/import/validate-data")
async def validate_import_data(
    activities: list[dict],
    current_user: Annotated[User, Depends(get_current_user)] = None,
):
    """
    Validate activity data using AI before import.

    Checks for:
    - Quantity reasonableness
    - Unit consistency
    - Missing required fields
    - Anomalies and outliers

    Returns validation results with suggested corrections.
    """
    if not activities:
        raise HTTPException(status_code=400, detail="No activities provided")

    if len(activities) > 100:
        raise HTTPException(
            status_code=400,
            detail="Validate up to 100 activities at a time"
        )

    validator = DataValidator()
    batch_result = validator.validate_batch(activities)

    return {
        "total_records": batch_result.total_records,
        "valid_count": batch_result.valid_count,
        "warning_count": batch_result.warning_count,
        "error_count": batch_result.error_count,
        "summary": batch_result.summary,
        "results": [
            {
                "index": i,
                "is_valid": r.is_valid,
                "issues": [
                    {
                        "field": issue.field,
                        "severity": issue.severity,
                        "message": issue.message,
                        "original_value": issue.original_value,
                        "suggested_value": issue.suggested_value,
                        "confidence": issue.confidence,
                    }
                    for issue in r.issues
                ],
                "corrected_data": r.corrected_data,
                "ai_notes": r.ai_notes,
            }
            for i, r in enumerate(batch_result.results)
        ],
    }


# ============================================================================
# GHG Data Collection Template Import (Multi-Sheet Excel)
# ============================================================================

from app.services.template_parser import TemplateParser


class TemplateImportPreview(BaseModel):
    """Preview of template parsing before import."""
    success: bool
    filename: str
    total_sheets: int
    processed_sheets: int
    total_activities: int
    by_scope: dict
    by_category: dict
    sheets: list[dict]
    errors: list[dict]
    warnings: list[str]


class TemplateImportResult(BaseModel):
    """Result of template import operation."""
    success: bool
    total_activities: int
    imported: int
    failed: int
    by_scope: dict
    by_category: dict
    errors: list[dict]
    warnings: list[str]
    import_batch_id: str | None = None


@router.post("/periods/{period_id}/import/template/preview", response_model=TemplateImportPreview)
async def preview_template_import(
    period_id: UUID,
    file: UploadFile = File(...),
    year: int = Query(default=None, description="Reporting year (default: current year)"),
    session: Annotated[AsyncSession, Depends(get_session)] = None,
    current_user: Annotated[User, Depends(get_current_user)] = None,
):
    """
    Preview GHG Data Collection Template before import.

    This endpoint parses the multi-sheet CLIMATRIX Excel template and returns
    a preview of all activities that would be imported. No data is saved.

    The template has specific sheets for each emission scope/category:
    - Scope1_Stationary: Fuel combustion in boilers, heaters, generators
    - Scope1_Mobile: Company vehicles, fleet fuel consumption
    - Scope1_Fugitive: Refrigerant leaks, A/C systems
    - Scope2_Electricity: Grid electricity consumption
    - Scope2_HeatSteam: District heating, purchased steam
    - Cat1_RawMaterials: Purchased raw materials (spend-based)
    - Cat1_Services: Purchased services (spend-based)
    - Cat4_9_Transport: Upstream/downstream transportation
    - Cat5_Waste: Waste disposal (landfill, recycling, etc.)
    - Cat6_BusinessTravel: Flights, hotels
    - Cat7_Commuting: Employee commuting
    - And more...
    """
    # Verify period access
    period_query = select(ReportingPeriod).where(
        ReportingPeriod.id == period_id,
        ReportingPeriod.organization_id == current_user.organization_id,
    )
    period_result = await session.execute(period_query)
    period = period_result.scalar_one_or_none()
    if not period:
        raise HTTPException(status_code=404, detail="Reporting period not found")

    # Check file type
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    filename_lower = file.filename.lower()
    if not filename_lower.endswith('.xlsx'):
        raise HTTPException(
            status_code=400,
            detail="Template must be an Excel file (.xlsx)"
        )

    # Read file content
    content = await file.read()

    # Parse template
    parser = TemplateParser(default_year=year)
    try:
        result = parser.parse(content, file.filename)
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Failed to parse template: {str(e)}"
        )

    return TemplateImportPreview(
        success=result.success,
        filename=result.filename,
        total_sheets=result.total_sheets,
        processed_sheets=result.processed_sheets,
        total_activities=result.total_activities,
        by_scope=result.by_scope,
        by_category=result.by_category,
        sheets=[
            {
                "sheet_name": s.sheet_name,
                "scope": s.scope,
                "category_code": s.category_code,
                "total_rows": s.total_rows,
                "parsed_rows": s.parsed_rows,
                "skipped_rows": s.skipped_rows,
                "errors": s.errors,
                "warnings": s.warnings,
            }
            for s in result.sheets
        ],
        errors=result.errors,
        warnings=result.warnings,
    )


@router.post("/periods/{period_id}/import/template", response_model=TemplateImportResult)
async def import_template(
    period_id: UUID,
    file: UploadFile = File(...),
    year: int = Query(default=None, description="Reporting year (default: current year)"),
    session: Annotated[AsyncSession, Depends(get_session)] = None,
    current_user: Annotated[User, Depends(get_current_user)] = None,
):
    """
    Import activities from GHG Data Collection Template.

    This endpoint parses the multi-sheet CLIMATRIX Excel template,
    calculates emissions for each activity, and saves them to the database.

    The template supports:
    - Scope 1: Stationary combustion, mobile combustion, fugitive emissions
    - Scope 2: Electricity, heat/steam
    - Scope 3: All 15 categories (purchased goods, transport, waste, travel, etc.)

    Both physical amounts (liters, kWh, km) and spend-based amounts (USD/ILS/EUR)
    are supported. Currency is automatically converted to USD.
    """
    # Verify period access
    period_query = select(ReportingPeriod).where(
        ReportingPeriod.id == period_id,
        ReportingPeriod.organization_id == current_user.organization_id,
    )
    period_result = await session.execute(period_query)
    period = period_result.scalar_one_or_none()
    if not period:
        raise HTTPException(status_code=404, detail="Reporting period not found")
    if period.is_locked:
        raise HTTPException(status_code=400, detail="Cannot import to locked period")

    # Get organization region
    org_query = select(Organization).where(Organization.id == current_user.organization_id)
    org_result = await session.execute(org_query)
    org = org_result.scalar_one_or_none()
    org_region = org.default_region if org and org.default_region else "Global"

    # Check file type
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    filename_lower = file.filename.lower()
    if not filename_lower.endswith('.xlsx'):
        raise HTTPException(
            status_code=400,
            detail="Template must be an Excel file (.xlsx)"
        )

    # Read file content
    content = await file.read()

    # Parse template
    parser = TemplateParser(default_year=year)
    try:
        parse_result = parser.parse(content, file.filename)
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Failed to parse template: {str(e)}"
        )

    if not parse_result.activities:
        raise HTTPException(
            status_code=400,
            detail="No activities found in template. Please check that data is entered correctly."
        )

    # Create ImportBatch to track this import
    import_batch = ImportBatch(
        organization_id=current_user.organization_id,
        reporting_period_id=period_id,
        file_name=file.filename,
        file_type='xlsx',
        file_size_bytes=len(content),
        total_rows=len(parse_result.activities),
        status=ImportBatchStatus.PROCESSING,
        uploaded_by=current_user.id,  # Required field!
    )
    session.add(import_batch)
    await session.flush()  # Get the batch ID

    # Process each parsed activity
    pipeline = CalculationPipeline(session)
    imported = 0
    failed = 0
    errors = []
    warnings = parse_result.warnings.copy()

    for activity_data in parse_result.activities:
        try:
            # Calculate emissions
            calc_result = await pipeline.calculate(ActivityInput(
                activity_key=activity_data.activity_key,
                quantity=activity_data.quantity,
                unit=activity_data.unit,
                scope=activity_data.scope,
                category_code=activity_data.category_code,
                region=org_region,
                year=year or datetime.now().year,
            ))

            # Parse activity date
            activity_date = date.today()
            if activity_data.activity_date:
                try:
                    activity_date = datetime.strptime(activity_data.activity_date, '%Y-%m-%d').date()
                except ValueError:
                    pass

            # Create activity
            activity = Activity(
                organization_id=current_user.organization_id,
                reporting_period_id=period_id,
                scope=activity_data.scope,
                category_code=activity_data.category_code,
                activity_key=activity_data.activity_key,
                description=activity_data.description,
                quantity=activity_data.quantity,
                unit=activity_data.unit,
                activity_date=activity_date,
                created_by=current_user.id,
                data_source=DataSource.IMPORT,
                import_batch_id=import_batch.id,
            )
            session.add(activity)
            await session.flush()

            # Create emission
            emission = Emission(
                activity_id=activity.id,
                emission_factor_id=calc_result.emission_factor_id,
                co2e_kg=calc_result.co2e_kg,
                co2_kg=calc_result.co2_kg,
                ch4_kg=calc_result.ch4_kg,
                n2o_kg=calc_result.n2o_kg,
                wtt_co2e_kg=calc_result.wtt_co2e_kg,
                converted_quantity=calc_result.converted_quantity,
                converted_unit=calc_result.converted_unit,
                formula=calc_result.formula,
                confidence=ConfidenceLevel(calc_result.confidence),
                resolution_strategy=calc_result.resolution_strategy,
            )
            session.add(emission)

            imported += 1

            # Collect activity warnings
            if activity_data.warnings:
                for w in activity_data.warnings:
                    warnings.append(f"Row {activity_data.source_row} ({activity_data.source_sheet}): {w}")

        except (FactorNotFoundError, UnitConversionError, CalculationError) as e:
            failed += 1
            errors.append({
                "sheet": activity_data.source_sheet,
                "row": activity_data.source_row,
                "activity_key": activity_data.activity_key,
                "error": str(e),
            })
        except Exception as e:
            failed += 1
            errors.append({
                "sheet": activity_data.source_sheet,
                "row": activity_data.source_row,
                "error": f"Unexpected error: {str(e)}",
            })

    # Update ImportBatch with results
    import_batch.successful_rows = imported
    import_batch.failed_rows = failed
    import_batch.status = ImportBatchStatus.COMPLETED if failed == 0 else ImportBatchStatus.PARTIAL
    import_batch.completed_at = datetime.utcnow()
    if errors:
        import_batch.row_errors = errors[:50]

    await session.commit()

    # Build summary by scope/category
    by_scope = {}
    by_category = {}
    for activity_data in parse_result.activities:
        scope_key = f"Scope {activity_data.scope}"
        if scope_key not in by_scope:
            by_scope[scope_key] = 0
        by_scope[scope_key] += 1

        if activity_data.category_code not in by_category:
            by_category[activity_data.category_code] = 0
        by_category[activity_data.category_code] += 1

    return TemplateImportResult(
        success=failed == 0,
        total_activities=len(parse_result.activities),
        imported=imported,
        failed=failed,
        by_scope=by_scope,
        by_category=by_category,
        errors=errors[:50],  # Limit errors
        warnings=warnings[:50],  # Limit warnings
        import_batch_id=str(import_batch.id),
    )


# ============================================================================
# Unified AI-Powered Import (handles ANY file type)
# ============================================================================

from app.services.ai.unified_import import UnifiedImportService, UnifiedImportPreview as UnifiedPreview


class UnifiedSheetPreview(BaseModel):
    """Preview of a single sheet from the unified import"""
    sheet_name: str
    detected_scope: int | None
    detected_category: str | None
    header_row: int
    total_rows: int
    columns: list[str]
    column_mappings: list[dict]
    sample_data: list[dict]
    activities_preview: list[dict]
    is_importable: bool
    skip_reason: str | None = None
    warnings: list[str] = []


class UnifiedImportPreviewResponse(BaseModel):
    """Complete preview of file for unified import"""
    success: bool
    file_name: str
    file_type: str
    total_sheets: int
    importable_sheets: int
    total_activities: int
    sheets: list[UnifiedSheetPreview]
    warnings: list[str] = []
    errors: list[str] = []


class UnifiedImportRequest(BaseModel):
    """Request to import with optional user-modified mappings"""
    user_mappings: dict[str, list[dict]] | None = None  # {sheet_name: [{column mappings}]}
    selected_sheets: list[str] | None = None  # Only import these sheets


class UnifiedImportResultResponse(BaseModel):
    """Result of unified import"""
    success: bool
    total_activities: int
    imported: int
    failed: int
    total_co2e_kg: float
    by_scope: dict[str, int]
    by_category: dict[str, int]
    errors: list[dict] = []
    warnings: list[str] = []


@router.post("/unified/preview", response_model=UnifiedImportPreviewResponse)
async def unified_import_preview(
    file: UploadFile = File(...),
    current_user: Annotated[User, Depends(get_current_user)] = None,
):
    """
    Analyze ANY file type using AI and return a preview.

    This endpoint:
    1. Detects file structure (CSV, single-sheet Excel, multi-sheet Excel)
    2. Finds header rows (even if not row 1)
    3. AI-maps columns to activity_keys
    4. Extracts sample activities for preview

    Works with:
    - Simple CSV files
    - Complex multi-sheet Excel templates (like iMDsoft with 19 sheets)
    - Files with headers not in row 1
    - Multi-language files (Hebrew, etc.)
    """
    # Read file content
    content = await file.read()
    filename = file.filename or "unknown.csv"

    # Process with unified import service
    service = UnifiedImportService()
    preview = service.analyze_file(content, filename)

    # Convert to response model
    sheet_previews = []
    for sheet in preview.sheets:
        sheet_previews.append(UnifiedSheetPreview(
            sheet_name=sheet.sheet_name,
            detected_scope=sheet.detected_scope,
            detected_category=sheet.detected_category,
            header_row=sheet.header_row,
            total_rows=sheet.total_rows,
            columns=sheet.columns,
            column_mappings=sheet.column_mappings,
            sample_data=sheet.sample_data,
            activities_preview=sheet.activities_preview,
            is_importable=sheet.is_importable,
            skip_reason=sheet.skip_reason,
            warnings=sheet.warnings,
        ))

    return UnifiedImportPreviewResponse(
        success=preview.success,
        file_name=preview.file_name,
        file_type=preview.file_type,
        total_sheets=preview.total_sheets,
        importable_sheets=preview.importable_sheets,
        total_activities=preview.total_activities,
        sheets=sheet_previews,
        warnings=preview.warnings,
        errors=preview.errors,
    )


@router.post("/unified/import/{period_id}", response_model=UnifiedImportResultResponse)
async def unified_import_activities(
    period_id: UUID,
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session),
    current_user: Annotated[User, Depends(get_current_user)] = None,
):
    """
    Import activities from ANY file type using AI-powered mapping.

    This endpoint:
    1. Analyzes file structure
    2. AI-maps columns to activity_keys
    3. Extracts and imports all activities
    4. Calculates emissions for each activity

    Supports optional user_mappings to override AI suggestions.
    """
    # Validate period exists and belongs to user's organization
    period = await session.get(ReportingPeriod, period_id)
    if not period or period.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Reporting period not found")

    # Read file content
    content = await file.read()
    filename = file.filename or "unknown.csv"

    # Get activities using unified import service
    service = UnifiedImportService()
    activities = service.import_with_mappings(content, filename)

    if not activities:
        return UnifiedImportResultResponse(
            success=False,
            total_activities=0,
            imported=0,
            failed=0,
            total_co2e_kg=0,
            by_scope={},
            by_category={},
            errors=[{"error": "No activities found in file"}],
        )

    # Get organization for region
    org = await session.get(Organization, current_user.organization_id)
    region = org.default_region if org else "Global"

    # Create calculation pipeline
    pipeline = CalculationPipeline(session)

    imported = 0
    failed = 0
    total_co2e = 0.0
    errors = []
    warnings = []
    by_scope = {}
    by_category = {}

    for activity_data in activities:
        try:
            # Create activity input
            activity_input = ActivityInput(
                scope=activity_data.scope,
                category_code=activity_data.category_code,
                activity_key=activity_data.activity_key,
                quantity=Decimal(str(activity_data.quantity)),
                unit=activity_data.unit,
                region=region,
            )

            # Calculate emission
            result = await pipeline.calculate(activity_input)

            # Create activity record
            activity = Activity(
                organization_id=current_user.organization_id,
                reporting_period_id=period_id,
                scope=activity_data.scope,
                category_code=activity_data.category_code,
                activity_key=activity_data.activity_key,
                description=activity_data.description or f"Imported: {activity_data.activity_key}",
                quantity=Decimal(str(activity_data.quantity)),
                unit=activity_data.unit,
                activity_date=date.fromisoformat(activity_data.activity_date) if activity_data.activity_date else date.today(),
                data_source=DataSource.IMPORT,
            )
            session.add(activity)
            await session.flush()

            # Create emission record
            emission = Emission(
                activity_id=activity.id,
                emission_factor_id=result.factor_used.id if result.factor_used else None,
                co2e_kg=result.co2e_kg,
                co2_kg=result.co2_kg,
                ch4_kg=result.ch4_kg,
                n2o_kg=result.n2o_kg,
                wtt_co2e_kg=result.wtt_co2e_kg,
                formula=result.formula,
                confidence=ConfidenceLevel.HIGH if activity_data.confidence >= 0.8 else ConfidenceLevel.MEDIUM,
            )
            session.add(emission)

            imported += 1
            total_co2e += float(result.co2e_kg)

            # Track by scope/category
            scope_key = f"Scope {activity_data.scope}"
            by_scope[scope_key] = by_scope.get(scope_key, 0) + 1
            by_category[activity_data.category_code] = by_category.get(activity_data.category_code, 0) + 1

            # Collect warnings
            if activity_data.warnings:
                for w in activity_data.warnings:
                    warnings.append(f"Row {activity_data.source_row} ({activity_data.source_sheet}): {w}")

        except (FactorNotFoundError, UnitConversionError, CalculationError) as e:
            failed += 1
            errors.append({
                "sheet": activity_data.source_sheet,
                "row": activity_data.source_row,
                "activity_key": activity_data.activity_key,
                "error": str(e),
            })
        except Exception as e:
            failed += 1
            errors.append({
                "sheet": activity_data.source_sheet,
                "row": activity_data.source_row,
                "error": f"Unexpected error: {str(e)}",
            })

    await session.commit()

    return UnifiedImportResultResponse(
        success=failed == 0,
        total_activities=len(activities),
        imported=imported,
        failed=failed,
        total_co2e_kg=total_co2e,
        by_scope=by_scope,
        by_category=by_category,
        errors=errors[:50],
        warnings=warnings[:50],
    )
