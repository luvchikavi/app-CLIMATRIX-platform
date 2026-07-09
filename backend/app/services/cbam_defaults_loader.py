"""
Loader for CBAM default emission values files.

Parses the EU Commission default-values file format (Excel published
13 Feb 2026, per CN code and origin country) — or a CSV with the same
columns — into plain dict rows ready for the `cbam_default_values` table.

Header matching is deliberately forgiving: the Commission files use long
descriptive headers; we match on keywords. Expected columns (any casing,
extra columns ignored):

- CN code            -> "cn code" / "cn_code" / "cn"
- Country of origin  -> "country" (ISO alpha-2 preferred; blank = any)
- Direct SEE         -> header containing "direct" (tCO2e/t)
- Indirect SEE       -> header containing "indirect"
- Total SEE          -> header containing "total" (falls back to
                        direct + indirect when absent)
- Year / version     -> header containing "year" / "version" (optional)

Kept free of DB imports so it is unit-testable without a session.
"""

from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

from app.services.cbam_screening import resolve_sector


class DefaultValuesParseError(ValueError):
    """Raised when the file cannot be interpreted as a default-values table."""


def _normalize_header(header: object) -> str:
    return str(header or "").strip().lower().replace("_", " ")


def _find_column(headers: list[str], *keywords: str) -> Optional[int]:
    """Return the index of the first header containing all keywords."""
    for idx, header in enumerate(headers):
        if all(kw in header for kw in keywords):
            return idx
    return None


def _to_decimal(value: object) -> Optional[Decimal]:
    if value is None:
        return None
    text = str(value).strip().replace(",", ".")
    if not text or text.lower() in {"n/a", "na", "-", "none"}:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _read_rows(path: Path) -> list[list[object]]:
    """Read raw rows from a .csv or .xlsx file."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        import csv

        with path.open(newline="", encoding="utf-8-sig") as fh:
            return [list(row) for row in csv.reader(fh)]
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()
        return rows
    raise DefaultValuesParseError(
        f"Unsupported file type '{suffix}' — expected .csv or .xlsx"
    )


def parse_default_values_file(path: str | Path) -> list[dict]:
    """
    Parse a Commission default-values file (xlsx or csv) into dict rows:

    {"cn_code", "country_code" (None = any), "sector", "direct_see",
     "indirect_see", "total_see", "dataset_year", "dataset_version",
     "product_description"}

    Rows without a CN code or without any SEE value are skipped.
    """
    path = Path(path)
    raw_rows = _read_rows(path)
    if not raw_rows:
        raise DefaultValuesParseError(f"{path.name} is empty")

    # Find the header row: first row containing a "cn" column.
    header_idx = None
    headers: list[str] = []
    for idx, row in enumerate(raw_rows[:20]):
        normalized = [_normalize_header(cell) for cell in row]
        if any(h == "cn" or "cn code" in h or h == "cn_code" for h in normalized):
            header_idx = idx
            headers = normalized
            break
    if header_idx is None:
        raise DefaultValuesParseError(
            f"Could not find a 'CN code' header row in {path.name}"
        )

    col_cn = _find_column(headers, "cn")
    col_country = _find_column(headers, "country")
    col_direct = _find_column(headers, "direct")
    col_indirect = _find_column(headers, "indirect")
    col_total = _find_column(headers, "total")
    col_year = _find_column(headers, "year")
    col_version = _find_column(headers, "version")
    col_desc = _find_column(headers, "description")
    if col_desc is None:
        col_desc = _find_column(headers, "good")

    def cell(row: list[object], idx: Optional[int]) -> object:
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    parsed: list[dict] = []
    for row in raw_rows[header_idx + 1 :]:
        cn_raw = str(cell(row, col_cn) or "").strip()
        cn_code = "".join(ch for ch in cn_raw if ch.isdigit())
        if not cn_code:
            continue

        direct = _to_decimal(cell(row, col_direct))
        indirect = _to_decimal(cell(row, col_indirect))
        total = _to_decimal(cell(row, col_total))
        if total is None and direct is not None:
            total = direct + (indirect or Decimal("0"))
        if total is None:
            continue

        country = str(cell(row, col_country) or "").strip().upper()
        country_code = country if len(country) == 2 else None

        year_raw = str(cell(row, col_year) or "").strip()
        dataset_year = int(year_raw) if year_raw.isdigit() else None

        version_raw = str(cell(row, col_version) or "").strip()

        parsed.append(
            {
                "cn_code": cn_code[:10],
                "country_code": country_code,
                "sector": resolve_sector(cn_code),
                "direct_see": direct if direct is not None else total,
                "indirect_see": indirect,
                "total_see": total,
                "dataset_year": dataset_year,
                "dataset_version": version_raw or None,
                "product_description": str(cell(row, col_desc) or "").strip()
                or f"CN {cn_code}",
            }
        )

    if not parsed:
        raise DefaultValuesParseError(
            f"No usable default-value rows found in {path.name}"
        )
    return parsed
