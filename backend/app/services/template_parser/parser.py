"""
Main template parser for CLIMATRIX GHG Data Collection Template.
"""
import io
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import BinaryIO

import openpyxl

from .models import ParsedActivity, ParseResult, SheetResult
from .sheet_config import get_sheet_config, get_all_sheet_configs, SheetConfig, create_auto_detect_config


class TemplateParser:
    """
    Parser for the CLIMATRIX GHG Data Collection Template.
    
    Handles multi-sheet Excel files with different structures per sheet.
    Converts all data to standardized ParsedActivity records.
    """
    
    def __init__(self, default_year: int = None):
        self.default_year = default_year or datetime.now().year
        self.currency_rates = {
            'ILS': 0.27,  # ILS to USD
            'EUR': 1.10,  # EUR to USD
            'GBP': 1.27,  # GBP to USD
            'USD': 1.0,
        }
    
    def parse(self, file_content: bytes, filename: str = "template.xlsx") -> ParseResult:
        """
        Parse a template Excel file.
        
        Args:
            file_content: Raw bytes of the Excel file
            filename: Original filename for tracking
            
        Returns:
            ParseResult with all parsed activities
        """
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        except Exception as e:
            return ParseResult(
                success=False,
                filename=filename,
                total_sheets=0,
                processed_sheets=0,
                total_activities=0,
                sheets=[],
                activities=[],
                errors=[{"error": f"Failed to open Excel file: {str(e)}"}],
            )
        
        all_activities = []
        sheet_results = []
        errors = []
        warnings = []
        
        configs = get_all_sheet_configs()
        
        for sheet_name in wb.sheetnames:
            # Skip non-data sheets (both legacy and generated template formats)
            skip_sheets = [
                'Instructions', 'ClientInfo', 'Suppliers', 'Lookups',  # Legacy
                'Introduction', 'Organization Info', 'Reference',  # Generated template
            ]
            if sheet_name in skip_sheets:
                continue

            config = get_sheet_config(sheet_name)

            # If no standard config, try auto-detection
            if not config:
                ws = wb[sheet_name]
                detected_config = self._auto_detect_sheet(ws, sheet_name)
                if detected_config:
                    config = detected_config
                    warnings.append(f"Auto-detected configuration for sheet: {sheet_name}")
                else:
                    warnings.append(f"No configuration for sheet: {sheet_name}")
                    continue

            try:
                sheet_result = self._parse_sheet(wb[sheet_name], config)
                sheet_results.append(sheet_result)
                all_activities.extend(sheet_result.activities)
            except Exception as e:
                errors.append({
                    "sheet": sheet_name,
                    "error": str(e),
                })
        
        wb.close()
        
        # Build summary
        by_scope = {}
        by_category = {}
        for activity in all_activities:
            # By scope
            if activity.scope not in by_scope:
                by_scope[activity.scope] = {"count": 0, "activities": []}
            by_scope[activity.scope]["count"] += 1
            # Add activity to the scope's activities list (frontend needs this)
            by_scope[activity.scope]["activities"].append({
                "scope": activity.scope,
                "category_code": activity.category_code,
                "activity_key": activity.activity_key,
                "description": activity.description,
                "quantity": float(activity.quantity) if activity.quantity else 0,
                "unit": activity.unit,
                "activity_date": activity.activity_date,
                "site": activity.site,
                "source_sheet": activity.source_sheet,
                "source_row": activity.source_row,
                "warnings": activity.warnings,
            })

            # By category
            if activity.category_code not in by_category:
                by_category[activity.category_code] = {"count": 0, "scope": activity.scope}
            by_category[activity.category_code]["count"] += 1
        
        return ParseResult(
            success=len(errors) == 0,
            filename=filename,
            total_sheets=len(wb.sheetnames),
            processed_sheets=len(sheet_results),
            total_activities=len(all_activities),
            sheets=sheet_results,
            activities=all_activities,
            by_scope=by_scope,
            by_category=by_category,
            errors=errors,
            warnings=warnings,
        )
    
    def _auto_detect_sheet(self, ws, sheet_name: str) -> SheetConfig | None:
        """
        Auto-detect sheet configuration by scanning for header row.

        Looks for rows that contain typical header keywords like
        'quantity', 'description', 'unit', etc.
        """
        header_keywords = ['quantity', 'amount', 'description', 'unit', 'date', 'scope', 'category']

        # Scan first 10 rows to find potential header row
        for row_num in range(1, min(11, ws.max_row + 1)):
            row_values = [str(cell.value or '').lower().strip() for cell in ws[row_num]]
            row_text = ' '.join(row_values)

            # Count how many header keywords are present
            keyword_count = sum(1 for kw in header_keywords if kw in row_text)

            # If we find at least 2 header keywords, assume this is the header row
            if keyword_count >= 2:
                headers = [str(cell.value).strip() if cell.value else '' for cell in ws[row_num]]
                config = create_auto_detect_config(sheet_name, headers)
                if config:
                    config.header_row = row_num
                    return config

        # Fallback: try row 1 as header
        if ws.max_row > 0:
            headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
            if any(headers):  # At least some headers
                config = create_auto_detect_config(sheet_name, headers)
                if config:
                    config.header_row = 1
                    return config

        return None

    def _parse_sheet(self, ws, config: SheetConfig) -> SheetResult:
        """Parse a single worksheet."""
        activities = []
        errors = []
        warnings = []
        skipped = 0
        
        # Get headers from header row
        headers = []
        for cell in ws[config.header_row]:
            headers.append(str(cell.value).strip() if cell.value else "")
        
        # Create column index map
        col_indices = {}
        for i, header in enumerate(headers):
            if header in config.column_map:
                col_indices[config.column_map[header]] = i
            else:
                # Try to find partial match
                for template_col, std_col in config.column_map.items():
                    if template_col in header:
                        col_indices[std_col] = i
                        break
        
        # Parse data rows
        total_rows = 0
        for row_num in range(config.header_row + 1, ws.max_row + 1):
            total_rows += 1
            row_values = [cell.value for cell in ws[row_num]]

            # Skip empty rows
            if not any(row_values):
                skipped += 1
                continue

            # Skip placeholder rows (e.g., "[Dropdown]", "[Paste/Type]")
            first_val = str(row_values[0] or '').strip()
            if first_val.startswith('[') and first_val.endswith(']'):
                skipped += 1
                continue
            
            try:
                activity = self._parse_row(
                    row_values, 
                    headers, 
                    col_indices, 
                    config, 
                    row_num
                )
                if activity:
                    activities.append(activity)
                else:
                    skipped += 1
            except Exception as e:
                errors.append({
                    "row": row_num,
                    "error": str(e),
                })
        
        return SheetResult(
            sheet_name=config.sheet_name,
            scope=config.scope,
            category_code=config.category_code,
            total_rows=total_rows,
            parsed_rows=len(activities),
            skipped_rows=skipped,
            activities=activities,
            errors=errors,
            warnings=warnings,
        )
    
    def _parse_row(
        self,
        row_values: list,
        headers: list,
        col_indices: dict,
        config: SheetConfig,
        row_num: int
    ) -> ParsedActivity | None:
        """Parse a single row into a ParsedActivity."""

        # Build row dict with header keys
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row_values):
                row_dict[header] = row_values[i]

        # Get quantity - check both Physical_Amount and Spend_Amount
        quantity = None
        unit = None
        is_spend = False

        # Check calc_type to determine which amount to use
        # Support both legacy ('Calc_Type') and generated template ('Method') column names
        calc_type = (row_dict.get('Calc_Type') or row_dict.get('Method') or '').lower().strip()

        if calc_type == 'spend' or config.is_spend_based:
            # For spend-based, use the quantity column (which has the amount)
            # and unit column (which has the currency)
            qty_idx = col_indices.get('quantity')
            if qty_idx is not None and qty_idx < len(row_values):
                qty_val = row_values[qty_idx]
                if qty_val:
                    quantity = self._parse_decimal(qty_val)

                    # Get currency from unit column
                    unit_idx = col_indices.get('unit')
                    currency = 'USD'
                    if unit_idx is not None and unit_idx < len(row_values):
                        currency = str(row_values[unit_idx] or 'USD').upper()

                    # Convert currency to USD if needed
                    if currency in self.currency_rates and currency != 'USD':
                        quantity = quantity * Decimal(str(self.currency_rates[currency]))

                    unit = 'USD'
                    is_spend = True

            # Also try legacy format with separate spend_amount column
            if quantity is None:
                spend_idx = col_indices.get('spend_amount')
                if spend_idx is not None and spend_idx < len(row_values):
                    spend_val = row_values[spend_idx]
                    if spend_val:
                        quantity = self._parse_decimal(spend_val)
                        # Convert currency to USD
                        currency_idx = col_indices.get('spend_currency')
                        currency = 'USD'
                        if currency_idx is not None and currency_idx < len(row_values):
                            currency = str(row_values[currency_idx] or 'USD').upper()

                        if currency in self.currency_rates and currency != 'USD':
                            quantity = quantity * Decimal(str(self.currency_rates[currency]))

                        unit = 'USD'
                        is_spend = True

        if quantity is None:
            # Use physical amount
            qty_idx = col_indices.get('quantity')
            if qty_idx is not None and qty_idx < len(row_values):
                qty_val = row_values[qty_idx]
                if qty_val:
                    quantity = self._parse_decimal(qty_val)

                    # Get unit
                    unit_idx = col_indices.get('unit')
                    if unit_idx is not None and unit_idx < len(row_values):
                        unit = str(row_values[unit_idx] or '').strip()

        # Special case: Cat7_Commuting - calculate from Employees × Working_Days × Avg_Distance × 2
        if config.sheet_name == 'Cat7_Commuting' and quantity is None:
            employees = self._parse_decimal(row_dict.get('Employees'))
            working_days = self._parse_decimal(row_dict.get('Working_Days'))
            avg_distance = self._parse_decimal(row_dict.get('Avg_Distance_km'))

            if employees and working_days and avg_distance:
                # Round trip = × 2
                quantity = employees * working_days * avg_distance * Decimal('2')
                unit = 'km'

        # Special case: 3.7 Commuting (new template) - same calculation
        if config.sheet_name == '3.7 Commuting' and quantity is None:
            employees = self._parse_decimal(row_dict.get('Number of Employees'))
            working_days = self._parse_decimal(row_dict.get('Working Days/Year') or row_dict.get('Working_Days'))
            avg_distance = self._parse_decimal(row_dict.get('Avg Distance (km one-way)') or row_dict.get('Avg_Distance_km'))

            if employees and working_days and avg_distance:
                # Round trip = × 2
                quantity = employees * working_days * avg_distance * Decimal('2')
                unit = 'km'

        # Special case: 3.4/3.9 Transport - calculate tonne-km from weight × distance
        if config.sheet_name in ['3.4 Upstream Transport', '3.9 Downstream Transport'] and quantity is None:
            weight = self._parse_decimal(row_dict.get('Weight (tonnes)'))
            distance = self._parse_decimal(row_dict.get('Distance (km)'))
            method = (row_dict.get('Method') or '').lower().strip()

            if method != 'spend' and weight and distance:
                quantity = weight * distance
                unit = 'tonne-km'
            elif method == 'spend':
                spend = self._parse_decimal(row_dict.get('Spend Amount'))
                if spend:
                    quantity = spend
                    unit = 'USD'

        # Special case: 3.6 Hotels - multiply nights × rooms
        if config.sheet_name == '3.6 Hotels':
            nights = self._parse_decimal(row_dict.get('Number of Nights'))
            rooms = self._parse_decimal(row_dict.get('Number of Rooms')) or Decimal('1')

            if nights:
                quantity = nights * rooms
                unit = 'nights'

        # Special case: Cat6_BusinessTravel - calculate distance from airports if not provided
        if config.sheet_name == 'Cat6_BusinessTravel' and quantity is None:
            from_airport = row_dict.get('From_Airport') or ''
            to_airport = row_dict.get('To_Airport') or ''
            trip_type = (row_dict.get('Trip_Type') or '').lower().strip()

            if from_airport and to_airport:
                # Calculate distance (use approximation table)
                distance = self._get_airport_distance(from_airport, to_airport)
                if distance:
                    # Round trip = × 2
                    if 'round' in trip_type:
                        quantity = Decimal(str(distance * 2))
                    else:
                        quantity = Decimal(str(distance))
                    unit = 'km'

        # Special case: 3.6 Flights (new template) - calculate distance from airports
        if config.sheet_name == '3.6 Flights' and quantity is None:
            from_airport = row_dict.get('From Airport') or ''
            to_airport = row_dict.get('To Airport') or ''
            trip_type = (row_dict.get('Trip Type') or '').lower().strip()
            num_trips = self._parse_decimal(row_dict.get('Number of Trips')) or Decimal('1')

            if from_airport and to_airport:
                # Calculate distance (use approximation table)
                distance = self._get_airport_distance(from_airport, to_airport)
                if distance:
                    # Round trip = × 2
                    if 'round' in trip_type:
                        quantity = Decimal(str(distance * 2)) * num_trips
                    else:
                        quantity = Decimal(str(distance)) * num_trips
                    unit = 'km'

        # Skip if no quantity
        if quantity is None or quantity == 0:
            return None
        
        # Get description
        desc_idx = col_indices.get('description')
        description = ""
        if desc_idx is not None and desc_idx < len(row_values):
            description = str(row_values[desc_idx] or '').strip()
        
        if not description:
            description = f"{config.sheet_name} - Row {row_num}"
        
        # Resolve activity_key
        if config.activity_key_resolver:
            activity_key, resolved_unit = config.activity_key_resolver(row_dict)
            if not unit:
                unit = resolved_unit
        else:
            activity_key = 'spend_other'
            if not unit:
                unit = 'USD'
        
        # Normalize unit
        unit = self._normalize_unit(unit)
        
        # Get site
        site_idx = col_indices.get('site')
        site = None
        if site_idx is not None and site_idx < len(row_values):
            site = str(row_values[site_idx] or '').strip() or None
        
        # Get year for activity_date
        year_idx = col_indices.get('year')
        year = self.default_year
        if year_idx is not None and year_idx < len(row_values):
            year_val = row_values[year_idx]
            if year_val:
                try:
                    year = int(year_val)
                except:
                    pass
        
        activity_date = f"{year}-06-30"  # Middle of year as default
        
        # Build warnings
        activity_warnings = []
        if is_spend:
            activity_warnings.append("Converted from spend-based calculation")
        
        return ParsedActivity(
            scope=config.scope,
            category_code=config.category_code,
            activity_key=activity_key,
            description=description,
            quantity=quantity,
            unit=unit,
            activity_date=activity_date,
            site=site,
            source_sheet=config.sheet_name,
            source_row=row_num,
            warnings=activity_warnings,
            raw_data=row_dict,
        )
    
    def _parse_decimal(self, value) -> Decimal | None:
        """Parse a value to Decimal."""
        if value is None:
            return None
        
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        
        try:
            # Remove commas and parse
            clean = str(value).replace(',', '').strip()
            return Decimal(clean)
        except (InvalidOperation, ValueError):
            return None
    
    def _normalize_unit(self, unit: str) -> str:
        """Normalize unit string."""
        if not unit:
            return ""

        unit = unit.lower().strip()

        # Unit normalization map
        normalizations = {
            'liters': 'liters',
            'liter': 'liters',
            'litres': 'liters',
            'litre': 'liters',
            'l': 'liters',
            'kwh': 'kWh',
            'mwh': 'MWh',
            'kw': 'kWh',
            'kg': 'kg',
            'kilogram': 'kg',
            'kilograms': 'kg',
            'gram': 'g',
            'grams': 'g',
            'g': 'g',
            'tonnes': 'tonnes',
            'tonne': 'tonnes',
            'ton': 'tonnes',
            'tons': 'tonnes',
            't': 'tonnes',
            'km': 'km',
            'kilometer': 'km',
            'kilometers': 'km',
            'kilometres': 'km',
            'm3': 'm3',
            'm³': 'm3',
            'cubic meter': 'm3',
            'cubic meters': 'm3',
            'nights': 'nights',
            'room-nights': 'nights',
            'room nights': 'nights',
            'usd': 'USD',
            '$': 'USD',
        }

        return normalizations.get(unit, unit)

    def _get_airport_distance(self, from_code, to_code) -> int | None:
        """
        Get distance between two airport codes using the full airport database.

        Uses Haversine formula for accurate great-circle distance calculation
        with 200+ airports worldwide.
        """
        # Handle potential non-string values (Excel might return dates)
        if not isinstance(from_code, str) or not isinstance(to_code, str):
            return None

        from_code = from_code.upper().strip()
        to_code = to_code.upper().strip()

        if not from_code or not to_code:
            return None

        # Use the full airport database for accurate distance calculation
        try:
            from app.data.airports import calculate_flight_distance
            distance = calculate_flight_distance(from_code, to_code)
            if distance:
                return int(distance)
        except ImportError:
            pass

        # Fallback: return None if airports not found
        return None
