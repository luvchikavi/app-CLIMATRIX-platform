"""
Generate Excel import template for CLIMATRIX bulk data upload.

This script creates a formatted Excel file with:
- Introduction sheet with instructions
- Organization info sheet
- Category-specific data entry sheets with dropdowns
- Reference sheet with valid values

Usage:
    python -m app.cli.generate_template
    python -m app.cli.generate_template --output /path/to/template.xlsx
"""

import typer
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from datetime import datetime

app = typer.Typer()

# =============================================================================
# REFERENCE DATA - Valid values for dropdowns
# =============================================================================

FUEL_TYPES_STATIONARY = [
    "Natural Gas",
    "Diesel",
    "Petrol",
    "LPG",
    "Coal",
    "Fuel Oil",
    "Kerosene",
    "Burning Oil",
]

UNITS_BY_FUEL = {
    "Natural Gas": ["m3", "kWh"],
    "Diesel": ["liters"],
    "Petrol": ["liters"],
    "LPG": ["liters", "kg"],
    "Coal": ["kg", "tonnes"],
    "Fuel Oil": ["liters"],
    "Kerosene": ["liters"],
    "Burning Oil": ["liters"],
}

VEHICLE_TYPES = [
    "Car",
    "Van",
    "LGV",
    "HGV",
    "Motorcycle",
    "Bus",
    "Taxi",
    "(Fuel Only)",
]

FUEL_TYPES_MOBILE = [
    "Petrol",
    "Diesel",
    "Hybrid",
    "Electric",
    "Plugin Hybrid",
]

MOBILE_UNITS = ["km", "liters"]

REFRIGERANT_TYPES = [
    "R-134a",
    "R-410A",
    "R-32",
    "R-404A",
    "R-407A",
    "R-407C",
    "R-407F",
    "R-417A",
    "R-422D",
    "R-507A",
    "R-508B",
    "HFC-23",
    "HFC-125",
    "HFC-143a",
    "HFC-152a",
    "HFC-227ea / FM-200",
    "R-123 / HCFC-123",
    "Halon-1211",
    "SF6",
    "NF3",
    "R-1234yf",
    "R-1234ze",
    "R-290 (Propane)",
    "R-600a (Isobutane)",
    "R-717 (Ammonia)",
    "R-744 (CO2)",
]

PROCESS_MATERIALS = [
    "Cement",
    "Clinker",
    "Quicklite (CaO)",
    "Dolomitic Lime",
    "Glass",
    "Iron & Steel (Integrated)",
    "Steel (EAF)",
    "Primary Aluminum",
    "Ammonia",
    "Nitric Acid",
    "Adipic Acid",
    "Hydrogen (SMR)",
    "Ethylene",
]

# Scope 2.1 - Electricity (separate from heat/cooling)
ELECTRICITY_TYPES = [
    "Grid Electricity (Location-based)",
    "Supplier Specific (Market-based)",
    "100% Renewable (Certified)",
]

# Scope 2.2 - Heat/Steam
HEAT_STEAM_TYPES = [
    "District Heating",
    "Steam",
]

# Scope 2.3 - Cooling
COOLING_TYPES = [
    "Chilled Water",
    "District Cooling",
]

# ALL countries from database (56 regions)
COUNTRIES = [
    "AE - UAE",
    "AR - Argentina",
    "AT - Austria",
    "AU - Australia",
    "BE - Belgium",
    "BR - Brazil",
    "CA - Canada",
    "CH - Switzerland",
    "CL - Chile",
    "CN - China",
    "CO - Colombia",
    "CZ - Czech Republic",
    "DE - Germany",
    "DK - Denmark",
    "EG - Egypt",
    "ES - Spain",
    "EU - European Union (Average)",
    "FI - Finland",
    "FR - France",
    "Global - World Average",
    "GR - Greece",
    "HK - Hong Kong",
    "HU - Hungary",
    "ID - Indonesia",
    "IE - Ireland",
    "IL - Israel",
    "IN - India",
    "IT - Italy",
    "JP - Japan",
    "KE - Kenya",
    "KR - South Korea",
    "MX - Mexico",
    "MY - Malaysia",
    "NG - Nigeria",
    "NL - Netherlands",
    "NO - Norway",
    "NZ - New Zealand",
    "PH - Philippines",
    "PL - Poland",
    "PT - Portugal",
    "RO - Romania",
    "RU - Russia",
    "SA - Saudi Arabia",
    "SE - Sweden",
    "SG - Singapore",
    "TH - Thailand",
    "TR - Turkey",
    "TW - Taiwan",
    "UK - United Kingdom",
    "US - United States (Average)",
    "US-CA - USA California",
    "US-MW - USA Midwest",
    "US-NY - USA New York",
    "US-TX - USA Texas (ERCOT)",
    "VN - Vietnam",
    "ZA - South Africa",
]

INDUSTRIES = [
    "Manufacturing",
    "Retail & Wholesale",
    "Financial Services",
    "Technology & Software",
    "Healthcare",
    "Construction",
    "Transportation & Logistics",
    "Food & Beverage",
    "Professional Services",
    "Real Estate",
    "Energy & Utilities",
    "Agriculture",
    "Education",
    "Hospitality",
    "Other",
]

EMPLOYEE_RANGES = [
    "1-10",
    "11-50",
    "51-100",
    "101-250",
    "251-500",
    "501-1000",
    "1001-5000",
    "5000+",
]

# Market methods for Scope 2 electricity
MARKET_METHODS = [
    "location_based",
    "market_based_supplier",
    "market_based_residual",
    "rec_ppa",
]

# Transport modes for Scope 3.4/3.9
TRANSPORT_MODES = [
    "Road - HGV",
    "Road - Van",
    "Road - Average",
    "Sea - Container",
    "Sea - Bulk",
    "Air Freight",
    "Rail Freight",
]

# Commuting transport modes for Scope 3.7
COMMUTING_MODES = [
    "Car - Petrol",
    "Car - Diesel",
    "Car - Hybrid",
    "Car - Electric",
    "Car - Average",
    "Bus",
    "Rail / Train",
    "Tram / Light Rail",
    "Motorcycle",
    "Bicycle",
    "Walking",
    "Work from Home",
]

# Leased asset building types for Scope 3.8
LEASED_BUILDING_TYPES = [
    "Office",
    "Retail",
    "Warehouse",
    "Industrial",
    "Data Center",
    "Mixed Use",
    "Other",
]

# Calculation methods
CALCULATION_METHODS = [
    "Physical",  # liters, kWh, km, kg, etc.
    "Spend",     # Currency amount - will be converted using fuel prices
]

# Currencies supported
CURRENCIES = [
    "USD",
    "EUR",
    "GBP",
    "ILS",
    "AUD",
    "CAD",
    "CHF",
    "CNY",
    "JPY",
    "Other",
]

# Scope 3 - Additional reference data
SCOPE3_METHODS = [
    "Physical",
    "Spend",
    "Supplier-Specific",
]

PURCHASED_GOODS_CATEGORIES = [
    "Raw Materials",
    "Office Supplies",
    "IT Equipment",
    "Packaging",
    "Chemicals",
    "Food & Beverages",
    "Textiles",
    "Other",
]

CAPITAL_GOODS_CATEGORIES = [
    "Vehicles",
    "IT Equipment",
    "Machinery",
    "Buildings",
    "Furniture",
    "HVAC",
    "Solar",
    "Other",
]

CAPITAL_GOODS_TYPES = [
    "Small Car",
    "Medium Car",
    "Large Car / SUV",
    "Van",
    "Truck / HGV",
    "Laptop",
    "Desktop",
    "Monitor",
    "Server",
    "Smartphone",
    "Tablet",
    "Printer",
    "Office",
    "Warehouse",
    "Retail",
    "Industrial",
    "HVAC System",
    "Solar PV",
    "Office Desk",
    "Office Chair",
]

WASTE_TYPES = [
    "Paper/Cardboard",
    "Plastics",
    "Glass",
    "Metals",
    "Organic/Food",
    "Wood",
    "Textile",
    "Mixed/General",
    "Electronic (WEEE)",
    "Construction & Demolition",
    "Hazardous",
    "Batteries",
]

TREATMENT_METHODS = [
    "Landfill",
    "Recycling",
    "Composting",
    "Incineration",
    "Incineration (Energy Recovery)",
    "Anaerobic Digestion",
]

CABIN_CLASSES = [
    "Economy",
    "Premium Economy",
    "Business",
    "First",
]

TRIP_TYPES = [
    "One-way",
    "Round-trip",
]

TRAVEL_TYPES = [
    "Rail",
    "Taxi",
    "Rental Car",
    "Bus",
]

PROCESSING_TYPES = [
    "Melting/Smelting",
    "Molding/Extrusion",
    "Assembly",
    "Weaving",
    "Milling/Refining",
    "Chemical Processing",
    "Other",
]

PRODUCT_CATEGORIES_USE = [
    "Vehicle",
    "Appliance",
    "Electronic",
    "Machinery",
    "Building",
    "Lighting",
    "Other",
]

DISPOSAL_METHODS = [
    "Recycling",
    "Landfill",
    "Incineration",
    "Incineration (Energy Recovery)",
    "Composting",
    "Anaerobic Digestion",
]

DOWNSTREAM_ASSET_TYPES = [
    "Office",
    "Warehouse",
    "Retail",
    "Industrial",
    "Data Center",
    "Residential",
    "Vehicle",
    "Equipment",
]

FRANCHISE_TYPES = [
    "Restaurant",
    "Fast Food",
    "Cafe/Coffee",
    "Retail Store",
    "Convenience Store",
    "Hotel/Hospitality",
    "Gym/Fitness",
    "Gas Station",
    "Office",
    "Service",
]

# =============================================================================
# STYLES
# =============================================================================

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F4E79")
SUBTITLE_FONT = Font(bold=True, size=12, color="1F4E79")
INSTRUCTION_FONT = Font(size=11)
EXAMPLE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
DROPDOWN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_introduction_sheet(ws):
    """Create the introduction/instructions sheet."""
    ws.title = "Introduction"

    # Set column widths
    ws.column_dimensions['A'].width = 100

    # Title
    ws['A1'] = "CLIMATRIX - GHG Emissions Data Import Template"
    ws['A1'].font = Font(bold=True, size=20, color="1F4E79")
    ws.row_dimensions[1].height = 30

    # Version info
    ws['A3'] = f"Template Version: 1.0 | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A3'].font = Font(italic=True, size=10, color="666666")

    # Welcome
    ws['A5'] = "Welcome to CLIMATRIX Data Import"
    ws['A5'].font = TITLE_FONT

    instructions = [
        "",
        "This template allows you to bulk import your emissions data into CLIMATRIX.",
        "Please follow the instructions below carefully to ensure successful data import.",
        "",
        "HOW TO USE THIS TEMPLATE",
        "=" * 50,
        "",
        "Step 1: Fill Organization Info (Optional but Recommended)",
        "   - Go to the 'Organization Info' sheet",
        "   - Enter your organization details",
        "   - This helps us configure your account correctly",
        "",
        "Step 2: Choose the Right Sheet for Your Data",
        "   SCOPE 1 - DIRECT EMISSIONS:",
        "   - 1.1 Stationary: Fuel burned in boilers, generators, heaters",
        "   - 1.2 Mobile: Company vehicles and fleet fuel",
        "   - 1.3 Fugitive: Refrigerant leaks and top-ups",
        "",
        "   SCOPE 2 - INDIRECT ENERGY EMISSIONS:",
        "   - 2.1 Electricity: Purchased grid electricity (56 countries, market-based supported)",
        "   - 2.2 Heat/Steam: District heating, industrial steam",
        "   - 2.3 Cooling: Chilled water, district cooling",
        "",
        "   SCOPE 3 - VALUE CHAIN EMISSIONS:",
        "   - 3.1 Purchased Goods: Raw materials, office supplies, packaging",
        "   - 3.2 Capital Goods: Vehicles, IT equipment, buildings, machinery",
        "   - 3.4 Upstream Transport: Freight & distribution (auto-distance from origin/destination)",
        "   - 3.5 Waste: Waste disposal by type and treatment method",
        "   - 3.6 Flights: Business air travel (auto-distance from airport codes)",
        "   - 3.6 Hotels: Hotel stays by country",
        "   - 3.6 Other Travel: Rail, taxi, rental car, bus",
        "   - 3.7 Commuting: Employee travel to work (Israel city/zone supported)",
        "   - 3.8 Leased Assets: Upstream leased buildings (tenant data supported)",
        "   - 3.9 Downstream Transport: Distribution of sold products",
        "   - 3.10 Processing: Processing of sold products by third parties",
        "   - 3.11 Use of Products: Emissions from use of sold products",
        "   - 3.12 End-of-Life: End-of-life treatment of sold products",
        "   - 3.13 Leased to Others: Downstream leased assets",
        "   - 3.14 Franchises: Emissions from franchise operations",
        "",
        "Step 3: Choose Calculation Method",
        "   - PHYSICAL: Enter actual consumption (liters, kWh, km, kg)",
        "   - SPEND: Enter cost amount (USD, EUR, GBP, etc.) - we convert using fuel prices",
        "",
        "Step 4: Enter Your Data",
        "   - YELLOW columns: Select from dropdown list",
        "   - WHITE columns: Paste or type your data",
        "   - GREEN rows: Example data (delete before uploading)",
        "",
        "Step 5: Review Your Data",
        "   - Check that all required columns are filled",
        "   - Ensure quantities are positive numbers",
        "   - Verify units/currencies match your calculation method",
        "",
        "Step 6: Save and Upload",
        "   - Save this file as .xlsx (Excel format)",
        "   - Upload to CLIMATRIX via Import page",
        "",
        "",
        "COLUMN COLOR GUIDE",
        "=" * 50,
        "",
        "   YELLOW background = Dropdown selection (click cell, select from list)",
        "   WHITE background = Free entry (paste or type your data)",
        "   GREEN row = Example data (for reference, delete before upload)",
        "",
        "",
        "TIPS FOR SUCCESS",
        "=" * 50,
        "",
        "   1. Use copy-paste for quantity and description columns",
        "   2. Select dropdown values AFTER pasting your quantities",
        "   3. Leave optional columns empty if you don't have the data",
        "   4. Date format: YYYY-MM-DD or YYYY-MM (e.g., 2024-01-15 or 2024-01)",
        "   5. Use decimal point (.) not comma (,) for numbers: 1234.56",
        "   6. Delete the green example row before uploading",
        "",
        "",
        "NEED HELP?",
        "=" * 50,
        "",
        "   - Check the 'Reference' sheet for all valid values",
        "   - Contact support@climatrix.com for assistance",
        "",
    ]

    row = 6
    for line in instructions:
        ws[f'A{row}'] = line
        if line.startswith("Step") or line.startswith("HOW") or line.startswith("COLUMN") or line.startswith("TIPS") or line.startswith("NEED"):
            ws[f'A{row}'].font = SUBTITLE_FONT
        elif line.startswith("   "):
            ws[f'A{row}'].font = INSTRUCTION_FONT
        row += 1


def create_org_info_sheet(ws):
    """Create the organization information sheet."""
    ws.title = "Organization Info"

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 50

    # Title
    ws['A1'] = "Organization Information"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:C1')

    ws['A2'] = "(Optional but recommended - helps configure your account)"
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:C2')

    # Headers
    headers = [
        ("Field", "Value", "Notes"),
    ]

    fields = [
        ("Organization Name", "", "Your company or organization name"),
        ("Industry Sector", "", "Select from dropdown"),
        ("Primary Country", "", "Main country of operations"),
        ("Number of Sites", "", "How many facilities/locations"),
        ("Number of Employees", "", "Select range from dropdown"),
        ("Reporting Period Start", "", "e.g., 2024-01-01"),
        ("Reporting Period End", "", "e.g., 2024-12-31"),
        ("Contact Email", "", "For import notifications"),
    ]

    # Write headers
    for col, header in enumerate(headers[0], 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # Write fields
    dropdown_fields = ["Industry Sector", "Primary Country", "Number of Employees"]
    for row_num, (field, value, notes) in enumerate(fields, 5):
        ws.cell(row=row_num, column=1, value=field).border = THIN_BORDER
        cell_b = ws.cell(row=row_num, column=2, value=value)
        cell_b.border = THIN_BORDER
        if field in dropdown_fields:
            cell_b.fill = DROPDOWN_FILL
        ws.cell(row=row_num, column=3, value=notes).font = Font(italic=True, color="666666")

    # Add dropdowns
    # Industry dropdown
    dv_industry = DataValidation(type="list", formula1='"' + ','.join(INDUSTRIES) + '"', allow_blank=True)
    dv_industry.error = "Please select from the list"
    dv_industry.errorTitle = "Invalid Industry"
    ws.add_data_validation(dv_industry)
    dv_industry.add(ws['B6'])

    # Country dropdown
    country_short = [c.split(' - ')[0] for c in COUNTRIES]
    dv_country = DataValidation(type="list", formula1='"' + ','.join(country_short) + '"', allow_blank=True)
    ws.add_data_validation(dv_country)
    dv_country.add(ws['B7'])

    # Employee range dropdown
    dv_emp = DataValidation(type="list", formula1='"' + ','.join(EMPLOYEE_RANGES) + '"', allow_blank=True)
    ws.add_data_validation(dv_emp)
    dv_emp.add(ws['B9'])


def create_scope_1_1_sheet(ws):
    """Create Scope 1.1 Stationary Combustion sheet."""
    ws.title = "1.1 Stationary"

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15

    # Title
    ws['A1'] = "Scope 1.1 - Stationary Combustion"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:G1')

    ws['A2'] = "Fuel burned in boilers, furnaces, generators, heaters. Use Physical (liters, kWh) OR Spend (currency)."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:G2')

    # Headers
    headers = ["Fuel Type", "Method", "Description", "Quantity/Amount", "Unit/Currency", "Date", "Site (Optional)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    # Add column type indicators
    indicators = ["[Dropdown]", "[Dropdown]", "[Paste/Type]", "[Paste/Type]", "[Dropdown]", "[Paste/Type]", "[Paste/Type]"]
    for col, ind in enumerate(indicators, 1):
        cell = ws.cell(row=5, column=col, value=ind)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center')

    # Example rows (green) - Physical AND Spend examples
    examples = [
        ["Natural Gas", "Physical", "Office heating - Q1 2024", "15000", "kWh", "2024-01", "HQ Building"],
        ["Diesel", "Physical", "Backup generator", "500", "liters", "2024-02", "Factory"],
        ["Natural Gas", "Spend", "Monthly gas bill", "2500", "USD", "2024-03", "Warehouse"],
    ]
    for row_offset, example in enumerate(examples):
        for col, ex in enumerate(example, 1):
            cell = ws.cell(row=6+row_offset, column=col, value=ex)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    # Data entry rows with formatting
    for row in range(9, 109):  # 100 data rows
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [1, 2, 5]:  # Dropdown columns
                cell.fill = DROPDOWN_FILL

    # Add dropdowns
    # Fuel type dropdown
    dv_fuel = DataValidation(type="list", formula1='"' + ','.join(FUEL_TYPES_STATIONARY) + '"', allow_blank=True)
    dv_fuel.error = "Please select a valid fuel type"
    dv_fuel.errorTitle = "Invalid Fuel Type"
    ws.add_data_validation(dv_fuel)
    dv_fuel.add('A6:A108')

    # Method dropdown (Physical vs Spend)
    dv_method = DataValidation(type="list", formula1='"' + ','.join(CALCULATION_METHODS) + '"', allow_blank=True)
    dv_method.error = "Select Physical or Spend"
    ws.add_data_validation(dv_method)
    dv_method.add('B6:B108')

    # Unit/Currency dropdown (combined)
    all_units = list(set(u for units in UNITS_BY_FUEL.values() for u in units)) + CURRENCIES
    dv_unit = DataValidation(type="list", formula1='"' + ','.join(all_units) + '"', allow_blank=True)
    dv_unit.error = "Please select a valid unit or currency"
    ws.add_data_validation(dv_unit)
    dv_unit.add('E6:E108')


def create_scope_1_2_sheet(ws):
    """Create Scope 1.2 Mobile Combustion sheet."""
    ws.title = "1.2 Mobile"

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12

    # Title
    ws['A1'] = "Scope 1.2 - Mobile Combustion"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:G1')

    ws['A2'] = "Company vehicles. Use Physical (km/liters) OR Spend (fuel cost). Select '(Fuel Only)' for fuel purchases without vehicle info."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:G2')

    # Headers
    headers = ["Vehicle Type", "Fuel Type", "Method", "Description", "Quantity/Amount", "Unit/Currency", "Date"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    # Add column type indicators
    indicators = ["[Dropdown]", "[Dropdown]", "[Dropdown]", "[Paste/Type]", "[Paste/Type]", "[Dropdown]", "[Paste/Type]"]
    for col, ind in enumerate(indicators, 1):
        cell = ws.cell(row=5, column=col, value=ind)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center')

    # Example rows (green) - Physical AND Spend
    examples = [
        ["Car", "Diesel", "Physical", "Sales team vehicle ABC-123", "25000", "km", "2024-01"],
        ["HGV", "Diesel", "Physical", "Delivery truck fleet", "85000", "km", "2024-01"],
        ["(Fuel Only)", "Diesel", "Physical", "Fleet fuel purchase", "5500", "liters", "2024-01"],
        ["(Fuel Only)", "Petrol", "Spend", "Monthly fuel cards", "8500", "USD", "2024-02"],
    ]
    for row_offset, example in enumerate(examples):
        for col, val in enumerate(example, 1):
            cell = ws.cell(row=6+row_offset, column=col, value=val)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    # Data entry rows
    for row in range(10, 110):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [1, 2, 3, 6]:  # Dropdown columns
                cell.fill = DROPDOWN_FILL

    # Add dropdowns
    dv_vehicle = DataValidation(type="list", formula1='"' + ','.join(VEHICLE_TYPES) + '"', allow_blank=True)
    ws.add_data_validation(dv_vehicle)
    dv_vehicle.add('A6:A109')

    dv_fuel = DataValidation(type="list", formula1='"' + ','.join(FUEL_TYPES_MOBILE) + '"', allow_blank=True)
    ws.add_data_validation(dv_fuel)
    dv_fuel.add('B6:B109')

    dv_method = DataValidation(type="list", formula1='"' + ','.join(CALCULATION_METHODS) + '"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add('C6:C109')

    # Unit/Currency dropdown
    mobile_units_currency = MOBILE_UNITS + CURRENCIES
    dv_unit = DataValidation(type="list", formula1='"' + ','.join(mobile_units_currency) + '"', allow_blank=True)
    ws.add_data_validation(dv_unit)
    dv_unit.add('F6:F109')


def create_scope_1_3_sheet(ws):
    """Create Scope 1.3 Fugitive Emissions sheet."""
    ws.title = "1.3 Fugitive"

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

    # Title
    ws['A1'] = "Scope 1.3 - Fugitive Emissions"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')

    ws['A2'] = "Refrigerant leaks, top-ups, equipment disposal. Use Physical (kg) OR Spend (purchase cost)."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:F2')

    # Headers
    headers = ["Refrigerant Type", "Method", "Description", "Quantity/Amount", "Unit/Currency", "Date"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    # Indicators
    indicators = ["[Dropdown]", "[Dropdown]", "[Paste/Type]", "[Paste/Type]", "[Dropdown]", "[Paste/Type]"]
    for col, ind in enumerate(indicators, 1):
        cell = ws.cell(row=5, column=col, value=ind)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center')

    # Example rows - Physical AND Spend
    examples = [
        ["R-410A", "Physical", "AC system refill - Building A", "3.5", "kg", "2024-03"],
        ["R-134a", "Spend", "Refrigerant purchase invoice", "450", "USD", "2024-04"],
    ]
    for row_offset, example in enumerate(examples):
        for col, ex in enumerate(example, 1):
            cell = ws.cell(row=6+row_offset, column=col, value=ex)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    # Data entry rows
    for row in range(8, 58):  # 50 rows
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [1, 2, 5]:
                cell.fill = DROPDOWN_FILL

    # Dropdowns
    dv_ref = DataValidation(type="list", formula1='"' + ','.join(REFRIGERANT_TYPES) + '"', allow_blank=True)
    ws.add_data_validation(dv_ref)
    dv_ref.add('A6:A57')

    dv_method = DataValidation(type="list", formula1='"' + ','.join(CALCULATION_METHODS) + '"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add('B6:B57')

    dv_unit = DataValidation(type="list", formula1='"kg,' + ','.join(CURRENCIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_unit)
    dv_unit.add('E6:E57')


def create_scope_1_4_sheet(ws):
    """Create Scope 1.4 Process Emissions sheet."""
    ws.title = "1.4 Process"

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

    # Title
    ws['A1'] = "Scope 1.4 - Process Emissions"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')

    ws['A2'] = "Industrial process emissions. Use Physical (tonnes produced) OR Spend (material purchase cost)."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:F2')

    # Headers
    headers = ["Material Type", "Method", "Description", "Quantity/Amount", "Unit/Currency", "Date"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    # Indicators
    indicators = ["[Dropdown]", "[Dropdown]", "[Paste/Type]", "[Paste/Type]", "[Dropdown]", "[Paste/Type]"]
    for col, ind in enumerate(indicators, 1):
        cell = ws.cell(row=5, column=col, value=ind)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center')

    # Examples - Physical AND Spend
    examples = [
        ["Cement", "Physical", "Portland cement production - Plant A", "25000", "tonnes", "2024-Q1"],
        ["Iron & Steel (Integrated)", "Spend", "Steel raw material purchase", "1500000", "USD", "2024-Q1"],
    ]
    for row_offset, example in enumerate(examples):
        for col, ex in enumerate(example, 1):
            cell = ws.cell(row=6+row_offset, column=col, value=ex)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    # Data rows
    for row in range(8, 38):  # 30 rows
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [1, 2, 5]:
                cell.fill = DROPDOWN_FILL

    # Dropdowns
    dv_mat = DataValidation(type="list", formula1='"' + ','.join(PROCESS_MATERIALS) + '"', allow_blank=True)
    ws.add_data_validation(dv_mat)
    dv_mat.add('A6:A37')

    dv_method = DataValidation(type="list", formula1='"' + ','.join(CALCULATION_METHODS) + '"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add('B6:B37')

    dv_unit = DataValidation(type="list", formula1='"tonnes,' + ','.join(CURRENCIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_unit)
    dv_unit.add('E6:E37')


def create_scope_2_1_sheet(ws):
    """Create Scope 2.1 Purchased Electricity sheet."""
    ws.title = "2.1 Electricity"

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12

    # Title
    ws['A1'] = "Scope 2.1 - Purchased Electricity"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:I1')

    ws['A2'] = "Grid electricity. Use Physical (kWh) OR Spend (electricity bill amount). Select country for correct grid factor. For market-based, select a power producer or method."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:I2')

    # Headers
    headers = [
        "Electricity Type", "Method", "Country/Region",
        "Power Producer", "Market Method",
        "Description", "Quantity/Amount", "Unit/Currency", "Date",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    # Add comments for new columns
    ws.cell(row=4, column=4).comment = Comment(
        "Optional. For Israel market-based: select the BDO power producer name. Leave blank for location-based.",
        "CLIMATRIX",
    )
    ws.cell(row=4, column=5).comment = Comment(
        "Optional. Choose: location_based (default), market_based_supplier (producer-specific), market_based_residual (residual mix), rec_ppa (REC/PPA instrument).",
        "CLIMATRIX",
    )

    # Indicators
    indicators = [
        "[Dropdown]", "[Dropdown]", "[Dropdown: 56]",
        "[Paste/Type]", "[Dropdown]",
        "[Paste/Type]", "[Paste/Type]", "[Dropdown]", "[Paste/Type]",
    ]
    for col, ind in enumerate(indicators, 1):
        cell = ws.cell(row=5, column=col, value=ind)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center')

    # Examples - Physical AND Spend, including market-based examples
    examples = [
        ["Grid Electricity (Location-based)", "Physical", "IL", "", "location_based", "Main office - Tel Aviv", "45000", "kWh", "2024-01"],
        ["Grid Electricity (Location-based)", "Physical", "US", "", "location_based", "US subsidiary", "28000", "kWh", "2024-01"],
        ["Grid Electricity (Location-based)", "Spend", "DE", "", "location_based", "Germany office - bill", "3500", "EUR", "2024-02"],
        ["Supplier Specific (Market-based)", "Physical", "IL", "OPC Energy", "market_based_supplier", "IL - OPC power producer", "22000", "kWh", "2024-01"],
        ["Supplier Specific (Market-based)", "Physical", "UK", "", "market_based_residual", "UK - residual mix", "15000", "kWh", "2024-01"],
        ["100% Renewable (Certified)", "Physical", "Global", "", "rec_ppa", "Data center - RECs", "180000", "kWh", "2024-Q1"],
    ]
    for row_offset, example in enumerate(examples):
        for col, val in enumerate(example, 1):
            cell = ws.cell(row=6+row_offset, column=col, value=val)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    # Data rows
    for row in range(12, 112):  # 100 rows
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [1, 2, 3, 5, 8]:  # Dropdown columns
                cell.fill = DROPDOWN_FILL

    # Dropdowns
    dv_type = DataValidation(type="list", formula1='"' + ','.join(ELECTRICITY_TYPES) + '"', allow_blank=True)
    ws.add_data_validation(dv_type)
    dv_type.add('A6:A111')

    dv_method = DataValidation(type="list", formula1='"' + ','.join(CALCULATION_METHODS) + '"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add('B6:B111')

    country_codes = [c.split(' - ')[0] for c in COUNTRIES]
    dv_country = DataValidation(type="list", formula1='"' + ','.join(country_codes) + '"', allow_blank=True)
    ws.add_data_validation(dv_country)
    dv_country.add('C6:C111')

    dv_market_method = DataValidation(type="list", formula1='"' + ','.join(MARKET_METHODS) + '"', allow_blank=True)
    dv_market_method.error = "Select a valid market method"
    dv_market_method.errorTitle = "Invalid Market Method"
    ws.add_data_validation(dv_market_method)
    dv_market_method.add('E6:E111')

    dv_unit = DataValidation(type="list", formula1='"kWh,' + ','.join(CURRENCIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_unit)
    dv_unit.add('H6:H111')


def create_scope_2_2_sheet(ws):
    """Create Scope 2.2 Purchased Heat/Steam sheet."""
    ws.title = "2.2 Heat-Steam"

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

    # Title
    ws['A1'] = "Scope 2.2 - Purchased Heat & Steam"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')

    ws['A2'] = "District heating and industrial steam. Use Physical (kWh) OR Spend (bill amount)."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:F2')

    # Headers
    headers = ["Energy Type", "Method", "Description", "Quantity/Amount", "Unit/Currency", "Date"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    # Indicators
    indicators = ["[Dropdown]", "[Dropdown]", "[Paste/Type]", "[Paste/Type]", "[Dropdown]", "[Paste/Type]"]
    for col, ind in enumerate(indicators, 1):
        cell = ws.cell(row=5, column=col, value=ind)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center')

    # Examples - Physical AND Spend
    examples = [
        ["District Heating", "Physical", "Office heating - Winter 2024", "35000", "kWh", "2024-Q1"],
        ["Steam", "Physical", "Manufacturing process steam", "18000", "kWh", "2024-01"],
        ["District Heating", "Spend", "Monthly heating bill", "1200", "EUR", "2024-02"],
    ]
    for row_offset, example in enumerate(examples):
        for col, val in enumerate(example, 1):
            cell = ws.cell(row=6+row_offset, column=col, value=val)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    # Data rows
    for row in range(9, 59):  # 50 rows
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [1, 2, 5]:
                cell.fill = DROPDOWN_FILL

    # Dropdowns
    dv_energy = DataValidation(type="list", formula1='"' + ','.join(HEAT_STEAM_TYPES) + '"', allow_blank=True)
    ws.add_data_validation(dv_energy)
    dv_energy.add('A6:A58')

    dv_method = DataValidation(type="list", formula1='"' + ','.join(CALCULATION_METHODS) + '"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add('B6:B58')

    dv_unit = DataValidation(type="list", formula1='"kWh,' + ','.join(CURRENCIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_unit)
    dv_unit.add('E6:E58')


def create_scope_2_3_sheet(ws):
    """Create Scope 2.3 Purchased Cooling sheet."""
    ws.title = "2.3 Cooling"

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

    # Title
    ws['A1'] = "Scope 2.3 - Purchased Cooling"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')

    ws['A2'] = "Chilled water and district cooling. Use Physical (kWh) OR Spend (bill amount)."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:F2')

    # Headers
    headers = ["Cooling Type", "Method", "Description", "Quantity/Amount", "Unit/Currency", "Date"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    # Indicators
    indicators = ["[Dropdown]", "[Dropdown]", "[Paste/Type]", "[Paste/Type]", "[Dropdown]", "[Paste/Type]"]
    for col, ind in enumerate(indicators, 1):
        cell = ws.cell(row=5, column=col, value=ind)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center')

    # Examples - Physical AND Spend
    examples = [
        ["Chilled Water", "Physical", "Data center cooling", "42000", "kWh", "2024-02"],
        ["District Cooling", "Physical", "Office cooling - Summer", "28000", "kWh", "2024-Q3"],
        ["District Cooling", "Spend", "Monthly cooling bill", "850", "USD", "2024-07"],
    ]
    for row_offset, example in enumerate(examples):
        for col, val in enumerate(example, 1):
            cell = ws.cell(row=6+row_offset, column=col, value=val)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    # Data rows
    for row in range(9, 59):  # 50 rows
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [1, 2, 5]:
                cell.fill = DROPDOWN_FILL

    # Dropdowns
    dv_cooling = DataValidation(type="list", formula1='"' + ','.join(COOLING_TYPES) + '"', allow_blank=True)
    ws.add_data_validation(dv_cooling)
    dv_cooling.add('A6:A58')

    dv_method = DataValidation(type="list", formula1='"' + ','.join(CALCULATION_METHODS) + '"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add('B6:B58')

    dv_unit = DataValidation(type="list", formula1='"kWh,' + ','.join(CURRENCIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_unit)
    dv_unit.add('E6:E58')


def create_scope_3_4_sheet(ws):
    """Create Scope 3.4 Upstream Transport sheet."""
    ws.title = "3.4 Upstream Transport"

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12

    # Title
    ws['A1'] = "Scope 3.4 - Upstream Transportation & Distribution"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:J1')

    ws['A2'] = "Transport of purchased goods. Enter origin/destination countries and distance will be auto-calculated if left blank."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:J2')

    ws['A3'] = "If Distance (km) is blank, origin + destination countries are used to look up a default route distance."
    ws['A3'].font = Font(italic=True, size=10, color="336699")
    ws.merge_cells('A3:J3')

    # Headers
    headers = [
        "Method", "Transport Mode", "Description", "Weight (tonnes)",
        "Origin Country", "Destination Country", "Distance (km)",
        "Spend Amount", "Currency", "Date",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    # Add comment on Distance (km) column
    ws.cell(row=5, column=7).comment = Comment(
        "Optional. If left blank the system will auto-calculate distance from the origin/destination countries using the built-in transport distance matrix.",
        "CLIMATRIX",
    )

    # Indicators
    indicators = [
        "[Dropdown]", "[Dropdown]", "[Paste/Type]", "[Paste/Type]",
        "[Dropdown]", "[Dropdown]", "[Optional]",
        "[Paste/Type]", "[Dropdown]", "[Paste/Type]",
    ]
    for col, ind in enumerate(indicators, 1):
        cell = ws.cell(row=6, column=col, value=ind)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center')

    # Examples
    examples = [
        ["Physical", "Sea - Container", "Raw materials from China", "120", "CN", "IL", "", "", "", "2024-Q1"],
        ["Physical", "Road - HGV", "Components from Germany", "25", "DE", "IL", "4200", "", "", "2024-02"],
        ["Physical", "Air Freight", "Urgent parts from US", "2.5", "US", "IL", "11000", "", "", "2024-03"],
        ["Spend", "Road - Average", "Local delivery (spend)", "", "", "", "", "15000", "USD", "2024-Q1"],
    ]
    for row_offset, example in enumerate(examples):
        for col, val in enumerate(example, 1):
            cell = ws.cell(row=7+row_offset, column=col, value=val)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    # Data rows
    for row in range(11, 111):  # 100 rows
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [1, 2, 5, 6, 9]:
                cell.fill = DROPDOWN_FILL

    # Dropdowns
    dv_method = DataValidation(type="list", formula1='"Physical,Spend"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add('A7:A110')

    dv_mode = DataValidation(type="list", formula1='"' + ','.join(TRANSPORT_MODES) + '"', allow_blank=True)
    ws.add_data_validation(dv_mode)
    dv_mode.add('B7:B110')

    country_codes = [c.split(' - ')[0] for c in COUNTRIES]
    dv_origin = DataValidation(type="list", formula1='"' + ','.join(country_codes) + '"', allow_blank=True)
    ws.add_data_validation(dv_origin)
    dv_origin.add('E7:E110')

    dv_dest = DataValidation(type="list", formula1='"' + ','.join(country_codes) + '"', allow_blank=True)
    ws.add_data_validation(dv_dest)
    dv_dest.add('F7:F110')

    dv_currency = DataValidation(type="list", formula1='"' + ','.join(CURRENCIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_currency)
    dv_currency.add('I7:I110')


def create_scope_3_7_sheet(ws):
    """Create Scope 3.7 Employee Commuting sheet."""
    ws.title = "3.7 Commuting"

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 12

    # Title
    ws['A1'] = "Scope 3.7 - Employee Commuting"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:K1')

    ws['A2'] = "Employee travel between home and work. Enter per group/mode or per site."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:K2')

    # Headers
    headers = [
        "Method", "Transport Mode", "Number of Employees",
        "Avg Distance One-Way (km)", "Working Days/Year",
        "Country", "City/Zone (Israel)",
        "% Remote Work", "Spend Amount", "Currency", "Site/Department",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    # Add comment for City/Zone column
    ws.cell(row=4, column=7).comment = Comment(
        "Optional. For Israel employees: enter city or zone name to use Israel-specific commuting factors (e.g., Tel Aviv, Haifa, Jerusalem, Gush Dan, Peripheral).",
        "CLIMATRIX",
    )

    # Indicators
    indicators = [
        "[Dropdown]", "[Dropdown]", "[Paste/Type]",
        "[Paste/Type]", "[Paste/Type]",
        "[Dropdown]", "[Paste/Type]",
        "[Paste/Type]", "[Paste/Type]", "[Dropdown]", "[Paste/Type]",
    ]
    for col, ind in enumerate(indicators, 1):
        cell = ws.cell(row=5, column=col, value=ind)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center')

    # Examples
    examples = [
        ["Physical", "Car - Petrol", "45", "25", "220", "IL", "Tel Aviv", "20", "", "", "R&D"],
        ["Physical", "Bus", "30", "18", "220", "IL", "Jerusalem", "10", "", "", "Operations"],
        ["Physical", "Rail / Train", "20", "35", "220", "UK", "", "15", "", "", "London Office"],
        ["Physical", "Work from Home", "15", "", "220", "IL", "Gush Dan", "100", "", "", "Support"],
        ["Spend", "Car - Average", "", "", "", "US", "", "", "45000", "USD", "US Subsidiary"],
    ]
    for row_offset, example in enumerate(examples):
        for col, val in enumerate(example, 1):
            cell = ws.cell(row=6+row_offset, column=col, value=val)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    # Data rows
    for row in range(11, 61):  # 50 rows
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [1, 2, 6, 10]:
                cell.fill = DROPDOWN_FILL

    # Dropdowns
    dv_method = DataValidation(type="list", formula1='"Physical,Spend"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add('A6:A60')

    dv_mode = DataValidation(type="list", formula1='"' + ','.join(COMMUTING_MODES) + '"', allow_blank=True)
    ws.add_data_validation(dv_mode)
    dv_mode.add('B6:B60')

    country_codes = [c.split(' - ')[0] for c in COUNTRIES]
    dv_country = DataValidation(type="list", formula1='"' + ','.join(country_codes) + '"', allow_blank=True)
    ws.add_data_validation(dv_country)
    dv_country.add('F6:F60')

    dv_currency = DataValidation(type="list", formula1='"' + ','.join(CURRENCIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_currency)
    dv_currency.add('J6:J60')


def create_scope_3_8_sheet(ws):
    """Create Scope 3.8 Upstream Leased Assets sheet."""
    ws.title = "3.8 Leased Assets"

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 25
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 12

    # Title
    ws['A1'] = "Scope 3.8 - Upstream Leased Assets"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:L1')

    ws['A2'] = "Emissions from leased buildings/assets not in Scopes 1 & 2. Use asset-specific data or spend-based method."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:L2')

    ws['A3'] = "If tenant-specific Scope 1 & 2 data is available, enter it in the dedicated columns for more accurate calculations."
    ws['A3'].font = Font(italic=True, size=10, color="336699")
    ws.merge_cells('A3:L3')

    # Headers
    headers = [
        "Method", "Building Type", "Description",
        "Floor Area (m2)", "Electricity (kWh)", "Gas (kWh)",
        "Country",
        "Tenant Scope 1 CO2e (kg)", "Tenant Scope 2 CO2e (kg)",
        "Spend Amount", "Currency", "Year",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    # Add comments for tenant columns
    ws.cell(row=5, column=8).comment = Comment(
        "Optional. If the tenant provides their own Scope 1 emissions data (direct emissions from fuel combustion, refrigerants, etc.), enter the total in kg CO2e.",
        "CLIMATRIX",
    )
    ws.cell(row=5, column=9).comment = Comment(
        "Optional. If the tenant provides their own Scope 2 emissions data (purchased electricity, heat, cooling), enter the total in kg CO2e.",
        "CLIMATRIX",
    )

    # Indicators
    indicators = [
        "[Dropdown]", "[Dropdown]", "[Paste/Type]",
        "[Paste/Type]", "[Paste/Type]", "[Paste/Type]",
        "[Dropdown]",
        "[Paste/Type]", "[Paste/Type]",
        "[Paste/Type]", "[Dropdown]", "[Paste/Type]",
    ]
    for col, ind in enumerate(indicators, 1):
        cell = ws.cell(row=6, column=col, value=ind)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center')

    # Examples
    examples = [
        ["Physical", "Office", "Leased HQ - Tel Aviv", "2500", "120000", "35000", "IL", "", "", "", "", "2024"],
        ["Physical", "Warehouse", "Distribution center", "5000", "85000", "", "IL", "4500", "18000", "", "", "2024"],
        ["Physical", "Data Center", "Cloud hosting facility", "800", "450000", "", "US", "12000", "95000", "", "", "2024"],
        ["Spend", "Office", "Small satellite office", "", "", "", "UK", "", "", "48000", "GBP", "2024"],
    ]
    for row_offset, example in enumerate(examples):
        for col, val in enumerate(example, 1):
            cell = ws.cell(row=7+row_offset, column=col, value=val)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    # Data rows
    for row in range(11, 61):  # 50 rows
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [1, 2, 7, 11]:
                cell.fill = DROPDOWN_FILL

    # Dropdowns
    dv_method = DataValidation(type="list", formula1='"Physical,Spend"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add('A7:A60')

    dv_building = DataValidation(type="list", formula1='"' + ','.join(LEASED_BUILDING_TYPES) + '"', allow_blank=True)
    ws.add_data_validation(dv_building)
    dv_building.add('B7:B60')

    country_codes = [c.split(' - ')[0] for c in COUNTRIES]
    dv_country = DataValidation(type="list", formula1='"' + ','.join(country_codes) + '"', allow_blank=True)
    ws.add_data_validation(dv_country)
    dv_country.add('G7:G60')

    dv_currency = DataValidation(type="list", formula1='"' + ','.join(CURRENCIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_currency)
    dv_currency.add('K7:K60')


def create_scope_3_1_sheet(ws):
    """Create Scope 3.1 Purchased Goods & Services sheet."""
    ws.title = "3.1 Purchased Goods"

    cols = {'A': 15, 'B': 18, 'C': 35, 'D': 15, 'E': 12, 'F': 15, 'G': 12, 'H': 22, 'I': 15, 'J': 12}
    for c, w in cols.items():
        ws.column_dimensions[c].width = w

    ws['A1'] = "Scope 3.1 - Purchased Goods & Services"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:J1')
    ws['A2'] = "Raw materials, components, office supplies, and services purchased. Use Physical, Spend, or Supplier-Specific."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:J2')

    headers = [
        "Method", "Sub-Category", "Description", "Quantity", "Unit",
        "Spend Amount", "Currency", "Supplier EF (kg CO2e/unit)", "Supplier Country", "Date",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    indicators = ["[Dropdown]", "[Paste/Type]", "[Paste/Type]", "[Paste/Type]", "[Dropdown]",
                  "[Paste/Type]", "[Dropdown]", "[Optional]", "[Optional]", "[Paste/Type]"]
    for col, ind in enumerate(indicators, 1):
        cell = ws.cell(row=5, column=col, value=ind)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center')

    examples = [
        ["Physical", "Steel", "Raw steel plates for production", "15000", "kg", "", "", "", "CN", "2024-Q1"],
        ["Spend", "Office Supplies", "Annual office supplies", "", "", "25000", "USD", "", "", "2024-01"],
        ["Supplier-Specific", "Plastic-PET", "PET bottles from supplier", "8000", "kg", "", "", "2.1", "IL", "2024-Q2"],
    ]
    for row_offset, example in enumerate(examples):
        for col, val in enumerate(example, 1):
            cell = ws.cell(row=6+row_offset, column=col, value=val)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    for row in range(9, 109):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [1, 5, 7]:
                cell.fill = DROPDOWN_FILL

    dv_method = DataValidation(type="list", formula1='"' + ','.join(SCOPE3_METHODS) + '"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add('A6:A108')

    dv_unit = DataValidation(type="list", formula1='"kg,tonnes,liters,m3,kWh,units,' + ','.join(CURRENCIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_unit)
    dv_unit.add('E6:E108')

    dv_currency = DataValidation(type="list", formula1='"' + ','.join(CURRENCIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_currency)
    dv_currency.add('G6:G108')


def create_scope_3_2_sheet(ws):
    """Create Scope 3.2 Capital Goods sheet."""
    ws.title = "3.2 Capital Goods"

    cols = {'A': 15, 'B': 18, 'C': 18, 'D': 35, 'E': 15, 'F': 12, 'G': 15, 'H': 12, 'I': 18, 'J': 12}
    for c, w in cols.items():
        ws.column_dimensions[c].width = w

    ws['A1'] = "Scope 3.2 - Capital Goods"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:J1')
    ws['A2'] = "Purchased capital equipment: vehicles, IT, buildings, machinery. Use Physical, Spend, or Supplier-Specific."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:J2')

    headers = [
        "Method", "Asset Category", "Asset Type", "Description", "Quantity", "Unit",
        "Spend Amount", "Currency", "Supplier EF", "Date",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    indicators = ["[Dropdown]", "[Dropdown]", "[Dropdown]", "[Paste/Type]", "[Paste/Type]", "[Dropdown]",
                  "[Paste/Type]", "[Dropdown]", "[Optional]", "[Paste/Type]"]
    for col, ind in enumerate(indicators, 1):
        cell = ws.cell(row=5, column=col, value=ind)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center')

    examples = [
        ["Physical", "IT Equipment", "Laptop", "50 laptops for employees", "50", "unit", "", "", "", "2024-01"],
        ["Physical", "Vehicles", "Medium Car", "Fleet vehicle purchase", "3", "unit", "", "", "", "2024-03"],
        ["Spend", "Machinery", "", "CNC machine purchase", "", "", "250000", "USD", "", "2024-Q2"],
    ]
    for row_offset, example in enumerate(examples):
        for col, val in enumerate(example, 1):
            cell = ws.cell(row=6+row_offset, column=col, value=val)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    for row in range(9, 59):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [1, 2, 3, 6, 8]:
                cell.fill = DROPDOWN_FILL

    dv_method = DataValidation(type="list", formula1='"' + ','.join(SCOPE3_METHODS) + '"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add('A6:A58')

    dv_cat = DataValidation(type="list", formula1='"' + ','.join(CAPITAL_GOODS_CATEGORIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_cat)
    dv_cat.add('B6:B58')

    dv_type = DataValidation(type="list", formula1='"' + ','.join(CAPITAL_GOODS_TYPES) + '"', allow_blank=True)
    ws.add_data_validation(dv_type)
    dv_type.add('C6:C58')

    dv_unit = DataValidation(type="list", formula1='"unit,m2,kW,' + ','.join(CURRENCIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_unit)
    dv_unit.add('F6:F58')

    dv_currency = DataValidation(type="list", formula1='"' + ','.join(CURRENCIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_currency)
    dv_currency.add('H6:H58')


def create_scope_3_5_sheet(ws):
    """Create Scope 3.5 Waste Generated in Operations sheet."""
    ws.title = "3.5 Waste"

    cols = {'A': 15, 'B': 20, 'C': 22, 'D': 35, 'E': 15, 'F': 12, 'G': 15, 'H': 12, 'I': 15, 'J': 12}
    for c, w in cols.items():
        ws.column_dimensions[c].width = w

    ws['A1'] = "Scope 3.5 - Waste Generated in Operations"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:J1')
    ws['A2'] = "Waste disposal and treatment. Enter waste type, treatment method, and weight."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:J2')

    headers = [
        "Method", "Waste Type", "Treatment Method", "Description", "Quantity", "Unit",
        "Spend Amount", "Currency", "Site", "Date",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    indicators = ["[Dropdown]", "[Dropdown]", "[Dropdown]", "[Paste/Type]", "[Paste/Type]", "[Dropdown]",
                  "[Paste/Type]", "[Dropdown]", "[Paste/Type]", "[Paste/Type]"]
    for col, ind in enumerate(indicators, 1):
        cell = ws.cell(row=5, column=col, value=ind)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center')

    examples = [
        ["Physical", "Paper/Cardboard", "Recycling", "Office paper waste", "2500", "kg", "", "", "HQ", "2024-Q1"],
        ["Physical", "Mixed/General", "Landfill", "General office waste", "8000", "kg", "", "", "HQ", "2024-Q1"],
        ["Spend", "Mixed/General", "Landfill", "Waste disposal contract", "", "", "12000", "USD", "", "2024"],
    ]
    for row_offset, example in enumerate(examples):
        for col, val in enumerate(example, 1):
            cell = ws.cell(row=6+row_offset, column=col, value=val)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    for row in range(9, 59):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [1, 2, 3, 6, 8]:
                cell.fill = DROPDOWN_FILL

    dv_method = DataValidation(type="list", formula1='"' + ','.join(SCOPE3_METHODS) + '"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add('A6:A58')

    dv_waste = DataValidation(type="list", formula1='"' + ','.join(WASTE_TYPES) + '"', allow_blank=True)
    ws.add_data_validation(dv_waste)
    dv_waste.add('B6:B58')

    dv_treatment = DataValidation(type="list", formula1='"' + ','.join(TREATMENT_METHODS) + '"', allow_blank=True)
    ws.add_data_validation(dv_treatment)
    dv_treatment.add('C6:C58')

    dv_unit = DataValidation(type="list", formula1='"kg,tonnes,m3,' + ','.join(CURRENCIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_unit)
    dv_unit.add('F6:F58')

    dv_currency = DataValidation(type="list", formula1='"' + ','.join(CURRENCIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_currency)
    dv_currency.add('H6:H58')


def create_scope_3_6_flights_sheet(ws):
    """Create Scope 3.6 Flights sheet."""
    ws.title = "3.6 Flights"

    cols = {'A': 15, 'B': 20, 'C': 20, 'D': 15, 'E': 12, 'F': 18, 'G': 12, 'H': 15, 'I': 12, 'J': 18, 'K': 12}
    for c, w in cols.items():
        ws.column_dimensions[c].width = w

    ws['A1'] = "Scope 3.6 - Business Travel: Flights"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:K1')
    ws['A2'] = "Business air travel. Enter IATA airport codes (e.g., TLV, LHR, JFK) - distance is auto-calculated."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:K2')
    ws['A3'] = "If the same airport appears as origin and destination (e.g., TLV→TLV), it will be treated as 0 emissions."
    ws['A3'].font = Font(italic=True, size=10, color="336699")
    ws.merge_cells('A3:K3')

    headers = [
        "Method", "Origin Airport", "Destination Airport", "Cabin Class", "Trip Type",
        "Passengers", "Number of Trips", "Spend Amount", "Currency", "Traveler Name", "Date",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    indicators = ["[Dropdown]", "[IATA Code]", "[IATA Code]", "[Dropdown]", "[Dropdown]",
                  "[Paste/Type]", "[Paste/Type]", "[Paste/Type]", "[Dropdown]", "[Paste/Type]", "[Paste/Type]"]
    for col, ind in enumerate(indicators, 1):
        cell = ws.cell(row=6, column=col, value=ind)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center')

    examples = [
        ["Physical", "TLV", "LHR", "Economy", "Round-trip", "1", "2", "", "", "John Smith", "2024-03"],
        ["Physical", "TLV", "JFK", "Business", "Round-trip", "1", "1", "", "", "Jane Doe", "2024-05"],
        ["Spend", "", "", "", "", "", "", "3500", "USD", "Team trip", "2024-06"],
    ]
    for row_offset, example in enumerate(examples):
        for col, val in enumerate(example, 1):
            cell = ws.cell(row=7+row_offset, column=col, value=val)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    for row in range(10, 110):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [1, 4, 5, 9]:
                cell.fill = DROPDOWN_FILL

    dv_method = DataValidation(type="list", formula1='"Physical,Spend"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add('A7:A109')

    dv_cabin = DataValidation(type="list", formula1='"' + ','.join(CABIN_CLASSES) + '"', allow_blank=True)
    ws.add_data_validation(dv_cabin)
    dv_cabin.add('D7:D109')

    dv_trip = DataValidation(type="list", formula1='"' + ','.join(TRIP_TYPES) + '"', allow_blank=True)
    ws.add_data_validation(dv_trip)
    dv_trip.add('E7:E109')

    dv_currency = DataValidation(type="list", formula1='"' + ','.join(CURRENCIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_currency)
    dv_currency.add('I7:I109')


def create_scope_3_6_hotels_sheet(ws):
    """Create Scope 3.6 Hotels sheet."""
    ws.title = "3.6 Hotels"

    cols = {'A': 15, 'B': 18, 'C': 18, 'D': 15, 'E': 15, 'F': 15, 'G': 12, 'H': 18, 'I': 12}
    for c, w in cols.items():
        ws.column_dimensions[c].width = w

    ws['A1'] = "Scope 3.6 - Business Travel: Hotels"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:I1')
    ws['A2'] = "Hotel stays during business travel. Enter number of nights, rooms, and country."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:I2')

    headers = [
        "Method", "Number of Nights", "Number of Rooms", "Country", "City",
        "Spend Amount", "Currency", "Traveler Name", "Date",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    indicators = ["[Dropdown]", "[Paste/Type]", "[Paste/Type]", "[Dropdown]", "[Paste/Type]",
                  "[Paste/Type]", "[Dropdown]", "[Paste/Type]", "[Paste/Type]"]
    for col, ind in enumerate(indicators, 1):
        cell = ws.cell(row=5, column=col, value=ind)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center')

    examples = [
        ["Physical", "3", "1", "UK", "London", "", "", "John Smith", "2024-03"],
        ["Physical", "5", "2", "US", "New York", "", "", "Team", "2024-06"],
        ["Spend", "", "", "", "", "850", "USD", "Jane Doe", "2024-04"],
    ]
    for row_offset, example in enumerate(examples):
        for col, val in enumerate(example, 1):
            cell = ws.cell(row=6+row_offset, column=col, value=val)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    for row in range(9, 59):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [1, 4, 7]:
                cell.fill = DROPDOWN_FILL

    dv_method = DataValidation(type="list", formula1='"Physical,Spend"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add('A6:A58')

    country_codes = [c.split(' - ')[0] for c in COUNTRIES]
    dv_country = DataValidation(type="list", formula1='"' + ','.join(country_codes) + '"', allow_blank=True)
    ws.add_data_validation(dv_country)
    dv_country.add('D6:D58')

    dv_currency = DataValidation(type="list", formula1='"' + ','.join(CURRENCIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_currency)
    dv_currency.add('G6:G58')


def create_scope_3_6_other_travel_sheet(ws):
    """Create Scope 3.6 Other Travel sheet."""
    ws.title = "3.6 Other Travel"

    cols = {'A': 18, 'B': 15, 'C': 15, 'D': 35, 'E': 15, 'F': 12, 'G': 12}
    for c, w in cols.items():
        ws.column_dimensions[c].width = w

    ws['A1'] = "Scope 3.6 - Business Travel: Other (Rail, Taxi, Rental Car, Bus)"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:G1')
    ws['A2'] = "Non-air business travel. Enter distance or spend amount."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:G2')

    headers = ["Travel Type", "Method", "Distance (km)", "Description", "Spend Amount", "Currency", "Date"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    indicators = ["[Dropdown]", "[Dropdown]", "[Paste/Type]", "[Paste/Type]", "[Paste/Type]", "[Dropdown]", "[Paste/Type]"]
    for col, ind in enumerate(indicators, 1):
        cell = ws.cell(row=5, column=col, value=ind)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center')

    examples = [
        ["Rail", "Physical", "450", "London to Manchester round trip", "", "", "2024-03"],
        ["Taxi", "Physical", "25", "Airport transfer", "", "", "2024-03"],
        ["Rental Car", "Spend", "", "Weekly rental", "850", "USD", "2024-04"],
    ]
    for row_offset, example in enumerate(examples):
        for col, val in enumerate(example, 1):
            cell = ws.cell(row=6+row_offset, column=col, value=val)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    for row in range(9, 59):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [1, 2, 6]:
                cell.fill = DROPDOWN_FILL

    dv_travel = DataValidation(type="list", formula1='"' + ','.join(TRAVEL_TYPES) + '"', allow_blank=True)
    ws.add_data_validation(dv_travel)
    dv_travel.add('A6:A58')

    dv_method = DataValidation(type="list", formula1='"Physical,Spend"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add('B6:B58')

    dv_currency = DataValidation(type="list", formula1='"' + ','.join(CURRENCIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_currency)
    dv_currency.add('F6:F58')


def create_scope_3_9_sheet(ws):
    """Create Scope 3.9 Downstream Transport sheet."""
    ws.title = "3.9 Downstream Transport"

    cols = {'A': 15, 'B': 35, 'C': 15, 'D': 18, 'E': 18, 'F': 15, 'G': 18, 'H': 15, 'I': 12, 'J': 18, 'K': 12}
    for c, w in cols.items():
        ws.column_dimensions[c].width = w

    ws['A1'] = "Scope 3.9 - Downstream Transportation & Distribution"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:K1')
    ws['A2'] = "Transport of sold products to customers. Same structure as 3.4 but for downstream."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:K2')

    headers = [
        "Method", "Description", "Weight (tonnes)",
        "Origin Country", "Destination Country", "Distance (km)",
        "Transport Mode", "Spend Amount", "Currency", "Customer/Region", "Date",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    indicators = ["[Dropdown]", "[Paste/Type]", "[Paste/Type]", "[Dropdown]", "[Dropdown]",
                  "[Optional]", "[Dropdown]", "[Paste/Type]", "[Dropdown]", "[Paste/Type]", "[Paste/Type]"]
    for col, ind in enumerate(indicators, 1):
        cell = ws.cell(row=5, column=col, value=ind)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center')

    examples = [
        ["Physical", "Products to EU market", "50", "IL", "DE", "", "Sea - Container", "", "", "EU Distribution", "2024-Q1"],
        ["Spend", "Local deliveries", "", "", "", "", "Road - Average", "8000", "USD", "Domestic", "2024-Q2"],
    ]
    for row_offset, example in enumerate(examples):
        for col, val in enumerate(example, 1):
            cell = ws.cell(row=6+row_offset, column=col, value=val)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    for row in range(8, 58):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [1, 4, 5, 7, 9]:
                cell.fill = DROPDOWN_FILL

    dv_method = DataValidation(type="list", formula1='"Physical,Spend"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add('A6:A57')

    country_codes = [c.split(' - ')[0] for c in COUNTRIES]
    dv_origin = DataValidation(type="list", formula1='"' + ','.join(country_codes) + '"', allow_blank=True)
    ws.add_data_validation(dv_origin)
    dv_origin.add('D6:D57')

    dv_dest = DataValidation(type="list", formula1='"' + ','.join(country_codes) + '"', allow_blank=True)
    ws.add_data_validation(dv_dest)
    dv_dest.add('E6:E57')

    dv_mode = DataValidation(type="list", formula1='"' + ','.join(TRANSPORT_MODES) + '"', allow_blank=True)
    ws.add_data_validation(dv_mode)
    dv_mode.add('G6:G57')

    dv_currency = DataValidation(type="list", formula1='"' + ','.join(CURRENCIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_currency)
    dv_currency.add('I6:I57')


def create_scope_3_10_sheet(ws):
    """Create Scope 3.10 Processing of Sold Products sheet."""
    ws.title = "3.10 Processing"

    cols = {'A': 15, 'B': 18, 'C': 18, 'D': 20, 'E': 35, 'F': 15, 'G': 12, 'H': 20, 'I': 15, 'J': 12, 'K': 12}
    for c, w in cols.items():
        ws.column_dimensions[c].width = w

    ws['A1'] = "Scope 3.10 - Processing of Sold Products"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:K1')
    ws['A2'] = "Emissions from downstream processing of intermediate products sold by your company."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:K2')

    headers = [
        "Method", "Product Type", "Product Category", "Processing Type", "Description",
        "Quantity Sold", "Unit", "Processing Energy (kWh)", "Supplier EF", "Currency", "Date",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    indicators = ["[Dropdown]", "[Paste/Type]", "[Paste/Type]", "[Dropdown]", "[Paste/Type]",
                  "[Paste/Type]", "[Dropdown]", "[Optional]", "[Optional]", "[Dropdown]", "[Paste/Type]"]
    for col, ind in enumerate(indicators, 1):
        cell = ws.cell(row=6, column=col, value=ind)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center')

    examples = [
        ["Physical", "Steel", "", "Melting/Smelting", "Steel plates for auto parts", "50000", "kg", "", "", "", "2024-Q1"],
        ["Spend", "", "", "", "Revenue from processed goods", "", "", "", "", "USD", "2024"],
    ]
    for row_offset, example in enumerate(examples):
        for col, val in enumerate(example, 1):
            cell = ws.cell(row=7+row_offset, column=col, value=val)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    for row in range(9, 59):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [1, 4, 7, 10]:
                cell.fill = DROPDOWN_FILL

    dv_method = DataValidation(type="list", formula1='"Physical,Spend,Site-Specific"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add('A7:A58')

    dv_process = DataValidation(type="list", formula1='"' + ','.join(PROCESSING_TYPES) + '"', allow_blank=True)
    ws.add_data_validation(dv_process)
    dv_process.add('D7:D58')

    dv_unit = DataValidation(type="list", formula1='"kg,tonnes,liters,units"', allow_blank=True)
    ws.add_data_validation(dv_unit)
    dv_unit.add('G7:G58')

    dv_currency = DataValidation(type="list", formula1='"' + ','.join(CURRENCIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_currency)
    dv_currency.add('J7:J58')


def create_scope_3_11_sheet(ws):
    """Create Scope 3.11 Use of Sold Products sheet."""
    ws.title = "3.11 Use of Products"

    cols = {'A': 15, 'B': 18, 'C': 18, 'D': 35, 'E': 15, 'F': 22, 'G': 15, 'H': 22, 'I': 15, 'J': 15, 'K': 12, 'L': 12}
    for c, w in cols.items():
        ws.column_dimensions[c].width = w

    ws['A1'] = "Scope 3.11 - Use of Sold Products"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:L1')
    ws['A2'] = "Emissions from the use of products sold by your company during their expected lifetime."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:L2')

    headers = [
        "Method", "Product Type", "Product Category", "Description", "Units Sold",
        "Lifetime Energy (kWh/unit)", "Lifetime (years)", "Lifetime Fuel (liters/unit)",
        "Fuel Type", "Energy Source", "Revenue", "Date",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    indicators = ["[Dropdown]", "[Dropdown]", "[Paste/Type]", "[Paste/Type]", "[Paste/Type]",
                  "[Paste/Type]", "[Paste/Type]", "[Paste/Type]", "[Dropdown]", "[Paste/Type]", "[Paste/Type]", "[Paste/Type]"]
    for col, ind in enumerate(indicators, 1):
        cell = ws.cell(row=6, column=col, value=ind)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center')

    examples = [
        ["Physical", "Appliance", "Air Conditioner", "Window AC units sold", "5000", "15000", "10", "", "", "Electricity", "", "2024"],
        ["Fuel", "Vehicle", "Truck", "Diesel trucks sold", "200", "", "15", "45000", "Diesel", "", "", "2024"],
    ]
    for row_offset, example in enumerate(examples):
        for col, val in enumerate(example, 1):
            cell = ws.cell(row=7+row_offset, column=col, value=val)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    for row in range(9, 59):
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [1, 2, 9]:
                cell.fill = DROPDOWN_FILL

    dv_method = DataValidation(type="list", formula1='"Physical,Fuel,Spend"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add('A7:A58')

    dv_type = DataValidation(type="list", formula1='"' + ','.join(PRODUCT_CATEGORIES_USE) + '"', allow_blank=True)
    ws.add_data_validation(dv_type)
    dv_type.add('B7:B58')

    dv_fuel = DataValidation(type="list", formula1='"Petrol,Diesel,Natural Gas,LPG,Electricity"', allow_blank=True)
    ws.add_data_validation(dv_fuel)
    dv_fuel.add('I7:I58')


def create_scope_3_12_sheet(ws):
    """Create Scope 3.12 End-of-Life Treatment of Sold Products sheet."""
    ws.title = "3.12 End-of-Life"

    cols = {'A': 15, 'B': 18, 'C': 18, 'D': 18, 'E': 35, 'F': 15, 'G': 15, 'H': 15, 'I': 15, 'J': 12, 'K': 12}
    for c, w in cols.items():
        ws.column_dimensions[c].width = w

    ws['A1'] = "Scope 3.12 - End-of-Life Treatment of Sold Products"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:K1')
    ws['A2'] = "Emissions from disposal and treatment of products sold by your company at end of life."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:K2')

    headers = [
        "Method", "Material Type", "Disposal Method", "Product Material", "Description",
        "Weight (kg)", "Units Sold", "Unit Weight (kg)", "Revenue", "Currency", "Date",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    indicators = ["[Dropdown]", "[Paste/Type]", "[Dropdown]", "[Paste/Type]", "[Paste/Type]",
                  "[Paste/Type]", "[Paste/Type]", "[Paste/Type]", "[Paste/Type]", "[Dropdown]", "[Paste/Type]"]
    for col, ind in enumerate(indicators, 1):
        cell = ws.cell(row=6, column=col, value=ind)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center')

    examples = [
        ["Physical", "Plastic", "Recycling", "PET packaging", "Product packaging waste", "25000", "", "", "", "", "2024"],
        ["Physical", "Electronic", "Recycling", "Circuit boards", "E-waste from sold devices", "5000", "10000", "0.5", "", "", "2024"],
    ]
    for row_offset, example in enumerate(examples):
        for col, val in enumerate(example, 1):
            cell = ws.cell(row=7+row_offset, column=col, value=val)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    for row in range(9, 59):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [1, 3, 10]:
                cell.fill = DROPDOWN_FILL

    dv_method = DataValidation(type="list", formula1='"Physical,Spend"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add('A7:A58')

    dv_disposal = DataValidation(type="list", formula1='"' + ','.join(DISPOSAL_METHODS) + '"', allow_blank=True)
    ws.add_data_validation(dv_disposal)
    dv_disposal.add('C7:C58')

    dv_currency = DataValidation(type="list", formula1='"' + ','.join(CURRENCIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_currency)
    dv_currency.add('J7:J58')


def create_scope_3_13_sheet(ws):
    """Create Scope 3.13 Downstream Leased Assets sheet."""
    ws.title = "3.13 Leased to Others"

    cols = {'A': 15, 'B': 18, 'C': 18, 'D': 35, 'E': 15, 'F': 18, 'G': 12, 'H': 15, 'I': 15, 'J': 15, 'K': 12, 'L': 18, 'M': 12}
    for c, w in cols.items():
        ws.column_dimensions[c].width = w

    ws['A1'] = "Scope 3.13 - Downstream Leased Assets"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:M1')
    ws['A2'] = "Assets owned by your company and leased to others. Enter energy data or use area-based estimates."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:M2')

    headers = [
        "Method", "Asset Type", "Building Type", "Description", "Energy Consumption",
        "Energy Unit", "Floor Area (m2)", "Number of Units", "Rental Income", "Currency",
        "Lessee", "Location", "Date",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    indicators = ["[Dropdown]", "[Dropdown]", "[Dropdown]", "[Paste/Type]", "[Paste/Type]",
                  "[Dropdown]", "[Paste/Type]", "[Paste/Type]", "[Paste/Type]", "[Dropdown]",
                  "[Paste/Type]", "[Paste/Type]", "[Paste/Type]"]
    for col, ind in enumerate(indicators, 1):
        cell = ws.cell(row=6, column=col, value=ind)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center')

    examples = [
        ["Physical", "Office", "Office", "Leased office space - Floor 3", "85000", "kWh", "1200", "", "", "", "Acme Corp", "Tel Aviv", "2024"],
        ["Spend", "Warehouse", "Warehouse", "Leased warehouse", "", "", "", "", "120000", "USD", "LogiCo", "Haifa", "2024"],
    ]
    for row_offset, example in enumerate(examples):
        for col, val in enumerate(example, 1):
            cell = ws.cell(row=7+row_offset, column=col, value=val)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    for row in range(9, 59):
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [1, 2, 3, 6, 10]:
                cell.fill = DROPDOWN_FILL

    dv_method = DataValidation(type="list", formula1='"Physical,Spend,Asset-Specific"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add('A7:A58')

    dv_asset = DataValidation(type="list", formula1='"' + ','.join(DOWNSTREAM_ASSET_TYPES) + '"', allow_blank=True)
    ws.add_data_validation(dv_asset)
    dv_asset.add('B7:B58')

    dv_building = DataValidation(type="list", formula1='"' + ','.join(LEASED_BUILDING_TYPES) + '"', allow_blank=True)
    ws.add_data_validation(dv_building)
    dv_building.add('C7:C58')

    dv_energy_unit = DataValidation(type="list", formula1='"kWh,MWh,m3,liters"', allow_blank=True)
    ws.add_data_validation(dv_energy_unit)
    dv_energy_unit.add('F7:F58')

    dv_currency = DataValidation(type="list", formula1='"' + ','.join(CURRENCIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_currency)
    dv_currency.add('J7:J58')


def create_scope_3_14_sheet(ws):
    """Create Scope 3.14 Franchises sheet."""
    ws.title = "3.14 Franchises"

    cols = {'A': 15, 'B': 18, 'C': 18, 'D': 35, 'E': 18, 'F': 18, 'G': 15, 'H': 15, 'I': 15, 'J': 12, 'K': 18, 'L': 18, 'M': 12}
    for c, w in cols.items():
        ws.column_dimensions[c].width = w

    ws['A1'] = "Scope 3.14 - Franchises"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:M1')
    ws['A2'] = "Emissions from franchise operations not in your Scope 1 & 2 (reported by franchisor)."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:M2')

    headers = [
        "Method", "Franchise Type", "Business Type", "Description", "Energy Consumption",
        "Fuel Consumption", "Number of Locations", "Floor Area (m2)", "Franchise Revenue",
        "Currency", "Franchisee Name", "Location", "Date",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    indicators = ["[Dropdown]", "[Dropdown]", "[Paste/Type]", "[Paste/Type]", "[Paste/Type]",
                  "[Paste/Type]", "[Paste/Type]", "[Paste/Type]", "[Paste/Type]", "[Dropdown]",
                  "[Paste/Type]", "[Paste/Type]", "[Paste/Type]"]
    for col, ind in enumerate(indicators, 1):
        cell = ws.cell(row=6, column=col, value=ind)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center')

    examples = [
        ["Physical", "Restaurant", "Fast Food", "Main St. franchise", "250000", "5000", "1", "350", "", "", "Franchisee A", "Tel Aviv", "2024"],
        ["Spend", "Retail Store", "", "Mall franchise", "", "", "3", "", "850000", "USD", "Franchisee B", "Haifa", "2024"],
    ]
    for row_offset, example in enumerate(examples):
        for col, val in enumerate(example, 1):
            cell = ws.cell(row=7+row_offset, column=col, value=val)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    for row in range(9, 39):
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [1, 2, 10]:
                cell.fill = DROPDOWN_FILL

    dv_method = DataValidation(type="list", formula1='"Physical,Spend,Franchise-Specific"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add('A7:A38')

    dv_type = DataValidation(type="list", formula1='"' + ','.join(FRANCHISE_TYPES) + '"', allow_blank=True)
    ws.add_data_validation(dv_type)
    dv_type.add('B7:B38')

    dv_currency = DataValidation(type="list", formula1='"' + ','.join(CURRENCIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_currency)
    dv_currency.add('J7:J38')


def create_reference_sheet(ws):
    """Create reference sheet with all valid values."""
    ws.title = "Reference"

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 25

    # Title
    ws['A1'] = "Reference - Valid Values for All Fields"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:E1')

    ws['A2'] = "This sheet shows all valid values accepted by the system. Use these exact values in your data."
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:E2')

    # Set column widths for all reference columns
    ref_col_widths = {
        'F': 30, 'G': 25, 'H': 20, 'I': 15, 'J': 15,
        'K': 25, 'L': 20, 'M': 15, 'N': 20, 'O': 25,
        'P': 22, 'Q': 18, 'R': 20, 'S': 18, 'T': 20,
    }
    for c, w in ref_col_widths.items():
        ws.column_dimensions[c].width = w

    headers = [
        "Fuels (1.1)",
        "Vehicles (1.2)",
        "Refrigerants (1.3)",
        "Countries (2.1)",
        "Electricity Types (2.1)",
        "Market Methods (2.1)",
        "Heat/Steam (2.2)",
        "Cooling (2.3)",
        "Transport Modes (3.4)",
        "Commuting Modes (3.7)",
        "Building Types (3.8)",
        "Calc Method",
        "Currencies",
        "Waste Types (3.5)",
        "Treatment Methods (3.5)",
        "Cabin Classes (3.6)",
        "Trip Types (3.6)",
        "Travel Types (3.6)",
        "Capital Asset Types (3.2)",
        "Franchise Types (3.14)",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # Values - all lists
    lists = [
        FUEL_TYPES_STATIONARY,
        VEHICLE_TYPES + FUEL_TYPES_MOBILE,
        REFRIGERANT_TYPES,
        COUNTRIES,
        ELECTRICITY_TYPES,
        MARKET_METHODS,
        HEAT_STEAM_TYPES,
        COOLING_TYPES,
        TRANSPORT_MODES,
        COMMUTING_MODES,
        LEASED_BUILDING_TYPES,
        CALCULATION_METHODS,
        CURRENCIES,
        WASTE_TYPES,
        TREATMENT_METHODS,
        CABIN_CLASSES,
        TRIP_TYPES,
        TRAVEL_TYPES,
        CAPITAL_GOODS_TYPES,
        FRANCHISE_TYPES,
    ]

    max_len = max(len(lst) for lst in lists)

    for row in range(5, 5 + max_len):
        for col, lst in enumerate(lists, 1):
            idx = row - 5
            if idx < len(lst):
                ws.cell(row=row, column=col, value=lst[idx]).border = THIN_BORDER


@app.command()
def generate(
    output: str = typer.Option(
        "climatrix_import_template.xlsx",
        "--output", "-o",
        help="Output file path"
    )
):
    """Generate the CLIMATRIX data import Excel template."""

    typer.echo("Generating CLIMATRIX import template...")

    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create all sheets
    create_introduction_sheet(wb.create_sheet())
    typer.echo("  + Introduction sheet")

    create_org_info_sheet(wb.create_sheet())
    typer.echo("  + Organization Info sheet")

    create_scope_1_1_sheet(wb.create_sheet())
    typer.echo("  + Scope 1.1 Stationary sheet")

    create_scope_1_2_sheet(wb.create_sheet())
    typer.echo("  + Scope 1.2 Mobile sheet")

    create_scope_1_3_sheet(wb.create_sheet())
    typer.echo("  + Scope 1.3 Fugitive sheet")

    create_scope_2_1_sheet(wb.create_sheet())
    typer.echo("  + Scope 2.1 Electricity sheet")

    create_scope_2_2_sheet(wb.create_sheet())
    typer.echo("  + Scope 2.2 Heat/Steam sheet")

    create_scope_2_3_sheet(wb.create_sheet())
    typer.echo("  + Scope 2.3 Cooling sheet")

    create_scope_3_1_sheet(wb.create_sheet())
    typer.echo("  + Scope 3.1 Purchased Goods sheet")

    create_scope_3_2_sheet(wb.create_sheet())
    typer.echo("  + Scope 3.2 Capital Goods sheet")

    create_scope_3_4_sheet(wb.create_sheet())
    typer.echo("  + Scope 3.4 Upstream Transport sheet")

    create_scope_3_5_sheet(wb.create_sheet())
    typer.echo("  + Scope 3.5 Waste sheet")

    create_scope_3_6_flights_sheet(wb.create_sheet())
    typer.echo("  + Scope 3.6 Flights sheet")

    create_scope_3_6_hotels_sheet(wb.create_sheet())
    typer.echo("  + Scope 3.6 Hotels sheet")

    create_scope_3_6_other_travel_sheet(wb.create_sheet())
    typer.echo("  + Scope 3.6 Other Travel sheet")

    create_scope_3_7_sheet(wb.create_sheet())
    typer.echo("  + Scope 3.7 Commuting sheet")

    create_scope_3_8_sheet(wb.create_sheet())
    typer.echo("  + Scope 3.8 Leased Assets sheet")

    create_scope_3_9_sheet(wb.create_sheet())
    typer.echo("  + Scope 3.9 Downstream Transport sheet")

    create_scope_3_10_sheet(wb.create_sheet())
    typer.echo("  + Scope 3.10 Processing sheet")

    create_scope_3_11_sheet(wb.create_sheet())
    typer.echo("  + Scope 3.11 Use of Products sheet")

    create_scope_3_12_sheet(wb.create_sheet())
    typer.echo("  + Scope 3.12 End-of-Life sheet")

    create_scope_3_13_sheet(wb.create_sheet())
    typer.echo("  + Scope 3.13 Downstream Leased Assets sheet")

    create_scope_3_14_sheet(wb.create_sheet())
    typer.echo("  + Scope 3.14 Franchises sheet")

    create_reference_sheet(wb.create_sheet())
    typer.echo("  + Reference sheet")

    # Save
    output_path = Path(output)
    wb.save(output_path)

    typer.echo(f"\n Template saved to: {output_path.absolute()}")
    typer.echo(f" File size: {output_path.stat().st_size / 1024:.1f} KB")


if __name__ == "__main__":
    app()
